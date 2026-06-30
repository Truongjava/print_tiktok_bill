"""
main_app.py — Desktop App TikTok Seller Automation + Bill Calculate
====================================================================
Chạy: python main_app.py
Đóng gói .exe: pyinstaller --onefile --windowed --name "TTS_Bill" main_app.py
"""
import os, sys, json, time, shutil, threading, queue
from datetime import datetime, timedelta
from pathlib import Path
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from playwright.sync_api import sync_playwright

sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'bill_calculate'))
try:
    from calculator import process_all
except ImportError:
    process_all = None  # Calculator không khả dụng (vd: khi chạy file .exe)

# ── Config ──
import os as _os
BASE_DIR = Path(__file__).parent
BILL_DIR = BASE_DIR / 'bill_calculate'
UPLOAD_DIR = BILL_DIR / 'uploads'

if getattr(sys, 'frozen', False):
    # Chạy file .exe — ưu tiên thư mục portable, fallback về system
    _portable_browsers = _os.path.join(_os.path.dirname(sys.executable), 'ms-playwright')
    if _os.path.isdir(_portable_browsers):
        _os.environ['PLAYWRIGHT_BROWSERS_PATH'] = _portable_browsers
    else:
        _os.environ.setdefault('PLAYWRIGHT_BROWSERS_PATH',
            _os.path.join(_os.path.expanduser('~'), 'AppData', 'Local', 'ms-playwright'))
    BASE_DIR = Path(sys.executable).parent
    BILL_DIR = BASE_DIR / 'bill_calculate'
    UPLOAD_DIR = BILL_DIR / 'uploads'
    DEFAULT_COOKIE = Path(sys._MEIPASS) / 'seller-vn.tiktok.com_25-06-2026.json'
    MASTER_DEFAULT = Path(sys._MEIPASS) / 'mã combo.xlsx'
    RETAIL_DEFAULT = Path(sys._MEIPASS) / 'sp bán lẻ.xlsx'
else:
    # Chạy source code
    _os.environ.setdefault('PLAYWRIGHT_BROWSERS_PATH',
        _os.path.join(_os.path.expanduser('~'), 'AppData', 'Local', 'ms-playwright'))
    DEFAULT_COOKIE = BASE_DIR / 'seller-vn.tiktok.com_25-06-2026.json'
    MASTER_DEFAULT = BASE_DIR / 'mã combo.xlsx'
    RETAIL_DEFAULT = BASE_DIR / 'sp bán lẻ.xlsx'

TARGET_URL = 'https://seller-vn.tiktok.com'
ORDERS_URL = 'https://seller-vn.tiktok.com/order?order_status%5B%5D=1&selected_sort=11&tab=to_ship&page_size=50'
# URL filter theo đơn vị vận chuyển
CARRIER_URLS = {
    'J&T':           ORDERS_URL + '&shipping_provider_id%5B%5D=6841743441349706241',
    'GHN':           ORDERS_URL + '&shipping_provider_id%5B%5D=7252807945006614278',
    'VietNam Post':  ORDERS_URL + '&shipping_provider_id%5B%5D=7062208235196909313',
    'Best Express':  ORDERS_URL + '&shipping_provider_id%5B%5D=7099655686241388293',
    'Viettel Post':  ORDERS_URL + '&shipping_provider_id%5B%5D=7155825439565416197',
    'J&T Cargo VN':  ORDERS_URL + '&shipping_provider_id%5B%5D=7581675938962736917',
}
BATCH_SIZE = 50
os.makedirs(UPLOAD_DIR, exist_ok=True)

# ============================================================
# AUTOMATION
# ============================================================
def run_automation(cookie_path, doc_type, output_dir, max_orders, log_cb, state_cb, stop_event=None,
                   existing_playwright=None, existing_browser=None, carrier=None, test_mode=False):
    """carrier: None (tất cả), 'J&T', hoặc 'GHN'. test_mode=True → chỉ chọn đơn, không in."""
    with open(cookie_path, 'r', encoding='utf-8') as f:
        cd = json.load(f)
    cookies_list = cd.get('cookies', cd if isinstance(cd, list) else [])
    pdf_files = []
    total_printed = 0
    target = max_orders if max_orders > 0 else 10**9
    use_select_all = (max_orders == 0)  # True = dùng nút "Chọn tất cả", False = tick từng đơn
    batch_num = 0
    carrier_label = f' [{carrier}]' if carrier else ''
    orders_url = CARRIER_URLS.get(carrier, ORDERS_URL)  # dùng URL filter nếu có carrier
    if stop_event is None:
        stop_event = threading.Event()  # fallback

    # ── Tái sử dụng browser nếu có, nếu không thì tạo mới ──
    if existing_browser and existing_playwright:
        playwright = existing_playwright
        browser = existing_browser
        # Dùng context hiện có (đã đăng nhập sẵn)
        context = browser.contexts[0] if browser.contexts else browser.new_context(
            viewport={'width': 1366, 'height': 768},
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/131.0.0.0 Safari/537.36',
            accept_downloads=True)
        # Tìm page có sẵn để F5 thay vì mở tab mới
        existing_pages = [p for p in context.pages if not p.is_closed()]
        if existing_pages:
            page = existing_pages[-1]  # dùng page cuối cùng (đang ở TikTok)
            log_cb('♻ Dùng lại browser — F5 trang đơn hàng...', 'info')
        else:
            page = context.new_page()
            log_cb('♻ Dùng lại browser — mở tab mới...', 'info')
        page.goto(orders_url, wait_until='networkidle', timeout=60000)
        page.wait_for_timeout(4000)
    else:
        playwright = sync_playwright().start()
        browser = playwright.chromium.launch(headless=False, args=['--disable-blink-features=AutomationControlled'])
        context = browser.new_context(viewport={'width': 1366, 'height': 768},
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/131.0.0.0 Safari/537.36',
            accept_downloads=True)
        pw_cookies = []
        for c in cookies_list:
            if not c.get('name') or not c.get('value'): continue
            pw = {'name': c['name'], 'value': c['value'], 'domain': c.get('domain', '.tiktok.com'), 'path': c.get('path', '/')}
            if c.get('expirationDate') and not c.get('session'): pw['expires'] = int(c['expirationDate'])
            if 'httpOnly' in c: pw['httpOnly'] = c['httpOnly']
            if 'secure' in c: pw['secure'] = c['secure']
            if c.get('sameSite'):
                pw['sameSite'] = {'strict':'Strict','lax':'Lax','no_restriction':'None','unspecified':'Lax'}.get(c['sameSite'],'Lax')
            pw_cookies.append(pw)
        context.add_cookies(pw_cookies)
        page = context.new_page()
        page.goto(TARGET_URL, wait_until='domcontentloaded', timeout=30000)
        page.wait_for_timeout(2000)

    try:
        while total_printed < target:
            if stop_event.is_set():
                log_cb('⏹ Đã dừng theo yêu cầu.', 'warn'); break
            batch_num += 1
            batch_target = min(BATCH_SIZE, target - total_printed)
            log_cb(f'▸ Batch {batch_num}{carrier_label}: chọn {batch_target} đơn (đã in {total_printed}/{target})', 'batch')

            # Vào trang đơn hàng
            state_cb('navigating', f'Batch {batch_num}: Đang tải danh sách đơn...')
            page.goto(orders_url, wait_until='networkidle', timeout=60000)
            page.wait_for_timeout(4000)

            # Đếm & chọn
            total_avail = page.evaluate("() => document.querySelectorAll('td.col-checkbox label.p-checkbox').length")
            if total_avail == 0:
                log_cb('  ✓ Hết đơn khả dụng — hoàn thành.', 'ok'); break

            if use_select_all:
                # ── Chế độ "Chọn tất cả": click checkbox header → bấm nút "Chọn tất cả X đơn" ──
                log_cb(f'  🖱 Đang chọn tất cả đơn hàng...', 'info')

                # B1: Click checkbox trong header — TikTok dùng SVG icon .p-checkbox-mask-icon
                header_clicked = False
                # Thứ tự ưu tiên: SVG icon → parent .p-checkbox → input[type=checkbox] → JS fallback
                for hdr_sel in [
                    'th svg.p-checkbox-mask-icon',           # SVG icon TikTok
                    'th .p-checkbox-mask-icon',               # class cha của SVG
                    'th .p-checkbox',                         # container checkbox header
                    'th.col-checkbox .p-checkbox',            # fallback cũ
                    'th.col-checkbox label.p-checkbox',       # fallback cũ
                    'th input[type="checkbox"]',               # input thật
                ]:
                    try:
                        hdr = page.locator(hdr_sel).first
                        if hdr.count() > 0 and hdr.is_visible(timeout=2000):
                            hdr.click(); header_clicked = True
                            log_cb(f'  ✓ Đã click checkbox header ({hdr_sel})', 'ok')
                            break
                    except: pass

                if not header_clicked:
                    # Fallback: JS click SVG icon hoặc container của nó
                    header_clicked = page.evaluate('''() => {
                        // Tìm SVG checkbox trong header
                        const svg = document.querySelector('th svg.p-checkbox-mask-icon, th .p-checkbox-mask-icon svg, th svg[class*="checkbox"]');
                        if (svg) {
                            // Click vào phần tử cha (có thể là div/span chứa SVG)
                            const parent = svg.closest('.p-checkbox') || svg.closest('label') || svg.closest('th');
                            if (parent) { parent.click(); return true; }
                            svg.dispatchEvent(new MouseEvent('click', {bubbles: true}));
                            return true;
                        }
                        // Fallback cũ
                        const cb = document.querySelector('th .p-checkbox, th.col-checkbox label, thead input[type="checkbox"]');
                        if (cb) { cb.click(); return true; }
                        return false;
                    }''')
                    if header_clicked:
                        log_cb('  ✓ Đã click checkbox header (JS fallback)', 'ok')

                page.wait_for_timeout(2500)

                # B2: Tìm và click nút "Chọn tất cả X đơn hàng" — TikTok dùng <span>
                select_all_btn = None
                for sel_text in ['Chọn tất cả', 'Select all', 'Chọn tất', 'Select All']:
                    for tag in ['span', 'button', 'div']:
                        try:
                            btns = page.locator(f'{tag}:has-text("{sel_text}")')
                            cnt = btns.count()
                            for j in range(cnt):
                                b = btns.nth(j)
                                if b.is_visible(timeout=800):
                                    btn_text = b.inner_text().strip()[:80]
                                    select_all_btn = b
                                    break
                        except: pass
                        if select_all_btn: break
                    if select_all_btn: break

                if select_all_btn:
                    select_all_btn.click(timeout=5000)
                    log_cb(f'  ✅ Đã bấm "Chọn tất cả"', 'ok')
                    page.wait_for_timeout(2000)
                    checked = page.evaluate("() => document.querySelectorAll('input[type=\"checkbox\"]:checked').length")
                    log_cb(f'  ✓ Đã chọn {checked} đơn hàng', 'ok')
                else:
                    # Fallback: tick từng đơn như cũ
                    log_cb('  ⚠ Không tìm thấy nút "Chọn tất cả" — fallback tick tay', 'warn')
                    to_select = min(batch_target, total_avail)
                    cbs = page.query_selector_all('td.col-checkbox label.p-checkbox')
                    checked = 0
                    for cb in cbs[:to_select]:
                        try: cb.click(); checked += 1; page.wait_for_timeout(120)
                        except: pass
                    page.wait_for_timeout(800)
                    log_cb(f'  ✓ Đã tick {checked}/{to_select} đơn (fallback)', 'ok')

                # Chọn tất cả → 1 lần duy nhất, force_stop sau khi in
                force_stop = True
                batch_target = 10**9  # vô hiệu hóa giới hạn batch
            else:
                # ── Chế độ tick từng đơn (có số lượng cụ thể) ──
                to_select = min(batch_target, total_avail)

                cbs = page.query_selector_all('td.col-checkbox label.p-checkbox')
                checked = 0
                for cb in cbs[:to_select]:
                    try: cb.click(); checked += 1; page.wait_for_timeout(120)
                    except: pass
                page.wait_for_timeout(800)

                # Nếu số đơn thực tế ít hơn batch_target → đây là batch cuối, in nốt rồi dừng
                force_stop = total_avail < batch_target
                log_cb(f'  ✓ Đã chọn {checked}/{batch_target} đơn', 'ok')

            if checked == 0:
                log_cb('  ✓ Hết đơn — hoàn thành.', 'ok'); break

            # ── Lưu ID các đơn hàng đã chọn vào Excel (phòng trường hợp tải lỗi) ──
            try:
                order_ids = page.evaluate('''() => {
                    const ids = [];
                    const rows = document.querySelectorAll('tr');
                    rows.forEach(row => {
                        const cb = row.querySelector('td.col-checkbox input[type="checkbox"]:checked');
                        if (cb) {
                            // Tìm ô chứa Order ID (thường là chuỗi số 15+ ký tự)
                            const cells = row.querySelectorAll('td');
                            cells.forEach(cell => {
                                const text = cell.textContent.trim();
                                if (/^\\d{15,}$/.test(text)) {
                                    ids.push(text);
                                }
                            });
                        }
                    });
                    return ids;
                }''')
                if order_ids:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, Alignment, Border, Side
                    order_file = str(Path(output_dir) / f'ID_don_hang_da_in_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
                    wb = Workbook(); ws = wb.active; ws.title = "Order IDs"
                    ws.cell(row=1, column=1, value="STT").font = Font(bold=True)
                    ws.cell(row=1, column=2, value="Order ID").font = Font(bold=True)
                    ws.cell(row=1, column=3, value="Batch").font = Font(bold=True)
                    for i, oid in enumerate(order_ids, 1):
                        ws.cell(row=i+1, column=1, value=i)
                        ws.cell(row=i+1, column=2, value=oid)
                        ws.cell(row=i+1, column=3, value=batch_num)
                    ws.column_dimensions['A'].width = 8
                    ws.column_dimensions['B'].width = 22
                    ws.column_dimensions['C'].width = 10
                    wb.save(order_file)
                    log_cb(f'  📋 Đã lưu {len(order_ids)} ID đơn hàng vào {Path(order_file).name}', 'ok')
            except Exception as e:
                log_cb(f'  ⚠ Không lưu được ID đơn hàng: {e}', 'warn')

            # ── TEST MODE: dừng sau khi chọn đơn, không in ──
            if test_mode:
                log_cb('  🧪 TEST MODE: Dừng tại bước chọn đơn — không in.', 'warn')
                try:
                    ss = str(Path(output_dir) / f'test_mode_batch{batch_num}.png')
                    page.screenshot(path=ss)
                    log_cb(f'  📸 Screenshot: {ss}', 'info')
                except: pass
                total_printed += checked
                log_cb(f'  📊 Đã chọn {checked} đơn (test mode — không in)', 'info')
                break

            # Click nút "Sắp xếp vận chuyển và in"
            state_cb('printing', f'Batch {batch_num}: Đang in...')
            ship_btn = None
            for btn_text in ['Sắp xếp vận chuyển và in', 'Arrange shipment and print', 'Sắp xếp vận chuyển']:
                try:
                    btn = page.locator(f'button:has-text("{btn_text}")').first
                    if btn.count() > 0 and btn.is_visible(timeout=3000):
                        ship_btn = btn; break
                except: pass
            if not ship_btn:
                # Fallback: tìm button chứa text "vận chuyển" và "in"
                all_btns = page.locator('button').all()
                for b in all_btns:
                    try:
                        txt = b.inner_text().strip().lower()
                        if 'vận chuyển' in txt and 'in' in txt:
                            ship_btn = b; break
                    except: pass
            if not ship_btn:
                log_cb('  ✗ KHÔNG TÌM THẤY nút "Sắp xếp vận chuyển và in"!', 'err')
                try:
                    ss = str(Path(output_dir) / f'debug_no_ship_btn_batch{batch_num}.png')
                    page.screenshot(path=ss); log_cb(f'  📸 Screenshot: {ss}', 'info')
                except: pass
                total_printed += checked; break
            ship_btn.click()
            log_cb('  ✓ Đã bấm "Sắp xếp vận chuyển và in"', 'ok')
            page.wait_for_timeout(3000)

            # ── Bước 1: Bấm "Tiếp theo" trong popup "Tính năng mới In nhãn" ──
            # NẾU CÓ "Tiếp theo" → lần chạy đầu: đi qua đầy đủ các bước popup
            # NẾU KHÔNG CÓ → lần chạy sau: TikTok bỏ qua hết, thẳng popup "Tải xuống tất cả"
            state_cb('printing', f'Batch {batch_num}: Đợi popup "Tiếp theo"...')
            tieptheo_btn = None
            for _ in range(10):
                for selector in ['button:has-text("Tiếp theo")', 'button:has-text("Next")']:
                    try:
                        btns = page.locator(selector)
                        cnt = btns.count()
                        for j in range(cnt):
                            b = btns.nth(j)
                            if b.is_visible(timeout=500):
                                tieptheo_btn = b; break
                    except: pass
                    if tieptheo_btn: break
                if tieptheo_btn: break
                page.wait_for_timeout(1000)

            if tieptheo_btn:
                # ── Lần đầu: có popup hướng dẫn → đi tuần tự ──
                tieptheo_btn.click(timeout=5000)
                log_cb('  ✓ Đã bấm "Tiếp theo"', 'ok')
                page.wait_for_timeout(3000)

                # Bước 2: Tick chọn Danh sách đóng gói + Danh sách lấy hàng
                state_cb('printing', f'Batch {batch_num}: Chọn loại chứng từ...')
                page.wait_for_timeout(2000)
                for doc_label in ['Danh sách đóng gói', 'Danh sách lấy hàng']:
                    try:
                        lbl = page.locator('label').filter(has_text=doc_label).first
                        if lbl.count() > 0:
                            inp = lbl.locator('input')
                            if inp.count() > 0 and not inp.is_checked():
                                lbl.click(); page.wait_for_timeout(300)
                                log_cb(f'  ✓ Đã tick: {doc_label}', 'ok')
                    except: pass
                page.wait_for_timeout(1000)

                # Bước 3: Bấm nút "In nhãn ngay sau khi vận chuyển"
                in_btn = None
                for btn_text in ['In nhãn ngay sau khi vận chuyển', 'In nhãn ngay', 'Print label immediately']:
                    try:
                        btn = page.locator(f'button:has-text("{btn_text}")').first
                        if btn.count() > 0 and btn.is_visible(timeout=2000):
                            in_btn = btn; break
                    except: pass
                if not in_btn:
                    for b in page.locator('button').all():
                        try:
                            txt = b.inner_text().strip().lower()
                            if 'in nhãn' in txt or 'in nhan' in txt:
                                in_btn = b; break
                        except: pass
                if in_btn:
                    in_btn.click(timeout=5000)
                    log_cb('  ✓ Đã bấm "In nhãn ngay"', 'ok')
                    page.wait_for_timeout(3000)
                else:
                    log_cb('  ⚠ Không tìm thấy nút "In nhãn ngay"', 'warn')
            else:
                # ── Lần sau: không có popup hướng dẫn → thẳng popup tải xuống ──
                log_cb('  ⏭ Không có popup "Tiếp theo" → đi thẳng bước tải xuống', 'info')

            # ── Bước 3: Popup "Tải xuống tất cả" ──
            state_cb('downloading', f'Batch {batch_num}: Đợi popup "Tải xuống tất cả"...')

            taixuong_btn = None
            for _ in range(30):
                btns = page.locator('button:has-text("Tải xuống tất cả"), button:has-text("Download all"), button:has-text("Tải xuống")')
                cnt = btns.count()
                for j in range(cnt):
                    b = btns.nth(j)
                    try:
                        if b.is_visible(timeout=500):
                            taixuong_btn = b; break
                    except: pass
                if taixuong_btn: break
                page.wait_for_timeout(1000)

            if not taixuong_btn:
                log_cb('  ✗ KHÔNG TÌM THẤY nút "Tải xuống tất cả tập tin" — kiểm tra popup!', 'err')
                try:
                    ss = str(Path(output_dir) / f'debug_no_taixuong_batch{batch_num}.png')
                    page.screenshot(path=ss); log_cb(f'  📸 Screenshot: {ss}', 'info')
                except: pass
                total_printed += checked; break

            log_cb('  📥 Đang bấm nút tải xuống...', 'info')
            # Dùng event listener để bắt TẤT CẢ download (nhiều file 1 lúc)
            downloaded_files = []
            def on_download(dl):
                carrier_prefix = carrier.replace(' ', '_').replace('&', 'n') + '_' if carrier else ''
                base_name = dl.suggested_filename or f'PDF_goc_TTS_batch{batch_num}_{len(downloaded_files)}_{datetime.now().strftime("%m-%d_%H-%M-%S")}.pdf'
                # File TikTok đã có sẵn ngày giờ, chỉ cần thêm tên hãng vào đầu
                suggested = carrier_prefix + base_name if carrier else base_name
                bp = str(Path(output_dir) / suggested)
                dl.save_as(bp)
                downloaded_files.append(bp)
                log_cb(f'  💾 Đã tải: {Path(bp).name}', 'ok')
            page.on('download', on_download)
            taixuong_btn.click(timeout=5000)
            log_cb('  ✓ Đã bấm "Tải xuống tất cả"', 'ok')
            # Đợi đủ lâu để tất cả download hoàn tất
            page.wait_for_timeout(8000)
            page.remove_listener('download', on_download)
            pdf_files.extend(downloaded_files)
            log_cb(f'  📥 Tổng cộng {len(downloaded_files)} file đã tải', 'info')
            if not downloaded_files:
                log_cb('  ✗ Không bắt được download nào!', 'err')
                total_printed += checked; break

            total_printed += checked
            log_cb(f'  📊 Tiến độ: {total_printed}/{target} đơn, {len(pdf_files)} file PDF', 'info')

            # Nếu đã in hết đơn khả dụng → dừng, không loop lại (tránh vòng lặp vô hạn khi có đơn mới)
            if force_stop:
                log_cb('  ✓ Đã in hết đơn hiện có — hoàn thành.', 'ok'); break

            if total_printed < target and total_avail > 0:
                page.wait_for_timeout(2000)

        # KHÔNG đóng browser — giữ lại cho lần chạy sau
        return pdf_files, playwright, browser
    except Exception as e:
        log_cb(f'  ✗ Lỗi: {e}', 'err')
        # KHÔNG đóng browser — giữ lại để xem lỗi
        return pdf_files, playwright, browser

# ============================================================
# CALCULATOR
# ============================================================
def run_calculator(pdf_paths, output_dir, master_path, retail_path, log_cb, carrier=''):
    if process_all is None:
        log_cb('✗ Calculator không khả dụng (thiếu module calculator)', 'err'); return []
    if not Path(master_path).exists(): log_cb(f'✗ Không tìm thấy master_data: {master_path}', 'err'); return []
    out_dir = str(output_dir)
    for p in pdf_paths:
        shutil.copy2(p, str(UPLOAD_DIR / Path(p).name))
    try:
        results = process_all(pdf_paths, out_dir, master_path, retail_path, carrier)
        for r in results:
            log_cb(f'  ✓ {r["rows"]} dòng | Qty={r["tong_qty"]} | Sold={r["tong_sold"]} | Promo={r["tong_promo"]}', 'ok')
            for key, fb in r['files'].items():
                src, dst = Path(fb), Path(out_dir) / Path(fb).name
                if src != dst and src.exists(): shutil.copy2(str(src), str(dst)); r['files'][key] = str(dst)
        return results
    except Exception as e: log_cb(f'  ✗ Lỗi: {e}', 'err'); return []

# ============================================================
# SCHEDULER
# ============================================================
class Scheduler:
    def __init__(self, callback, log_cb, state_cb):
        self.callback = callback; self.log = log_cb; self.set_state = state_cb
        self.running = False; self.mode = 'once'; self.interval_hours = 1
        self.daily_times = []; self.next_run = None; self.last_run = None
        self._stop = threading.Event(); self._thread = None

    def configure(self, mode, interval_hours=1, daily_times=None):
        self.mode = mode; self.interval_hours = interval_hours; self.daily_times = daily_times or []

    def start(self):
        if self._thread and self._thread.is_alive(): return
        self.running = True; self._stop.clear()
        self._thread = threading.Thread(target=self._loop, daemon=True); self._thread.start()

    def stop(self):
        self.running = False; self._stop.set(); self.set_state('idle', '⏹ Đã dừng')

    def _loop(self):
        while self.running and not self._stop.is_set():
            now = datetime.now()
            if self.mode == 'once': self._run_job(); self.running = False; break
            elif self.mode == 'interval':
                self.next_run = (self.last_run or now) + timedelta(hours=self.interval_hours)
                if now >= self.next_run: self._run_job()
                else:
                    w = (self.next_run - now).total_seconds()
                    self._countdown(w)
                    for _ in range(int(min(w, 60))):
                        if self._stop.is_set(): return
                        time.sleep(1)
            elif self.mode == 'daily':
                runs = []
                for t in self.daily_times:
                    try:
                        h, m = t.strip().split(':'); rt = now.replace(hour=int(h), minute=int(m), second=0, microsecond=0)
                        if rt <= now: rt += timedelta(days=1)
                        runs.append(rt)
                    except: pass
                if runs:
                    self.next_run = min(runs); w = (self.next_run - now).total_seconds()
                    if w <= 1: self._run_job()
                    else:
                        self._countdown(w)
                        for _ in range(int(min(w, 60))):
                            if self._stop.is_set(): return; time.sleep(1)
                else: time.sleep(60)
            time.sleep(1)

    def _run_job(self):
        self.set_state('running', '🔄 Đang chạy...'); start = datetime.now()
        try: self.callback()
        except Exception as e: self.log(f'✗ Job lỗi: {e}', 'err')
        self.last_run = datetime.now()
        self.log(f'🏁 Hoàn thành ({int((datetime.now()-start).total_seconds())}s)', 'ok')

    def _countdown(self, seconds):
        h, m = int(seconds//3600), int((seconds%3600)//60)
        self.set_state('waiting', f'⏳ Chạy tiếp sau {h}h{m:02d}')

# ============================================================
# AUTOMATION WORKER — dedicated thread cho Playwright
# ============================================================
class AutomationWorker:
    """Chạy tất cả automation trên 1 thread duy nhất để tránh lỗi thread-safety của Playwright.
    Browser được tạo 1 lần và tái sử dụng cho mọi job về sau."""

    def __init__(self):
        self._queue = queue.Queue()
        self._playwright = None
        self._browser = None
        self._thread = None
        self._running = False
        self._job_stop = threading.Event()

    def start(self):
        """Khởi động worker thread."""
        self._running = True
        self._thread = threading.Thread(target=self._loop, daemon=True)
        self._thread.start()

    def run_job(self, job_fn):
        """Gửi 1 job vào queue. job_fn(pw, browser, stop_event) -> result được gọi trên worker thread."""
        self._queue.put((job_fn,))

    def stop_job(self):
        """Ra hiệu dừng job hiện tại (không đóng browser)."""
        self._job_stop.set()

    def shutdown(self):
        """Dừng worker thread và đóng browser."""
        self._running = False
        self._queue.put(None)  # sentinel để thoát _loop

    def _loop(self):
        """Vòng lặp chính của worker thread."""
        while self._running:
            item = self._queue.get()
            if item is None:
                break
            (job_fn,) = item
            self._job_stop.clear()
            try:
                result = job_fn(self._playwright, self._browser, self._job_stop)
                # Lưu browser/playwright nếu job vừa tạo mới
                if isinstance(result, dict):
                    if 'playwright' in result:
                        self._playwright = result['playwright']
                    if 'browser' in result:
                        self._browser = result['browser']
            except Exception:
                pass  # Lỗi đã được job_fn tự xử lý và log

        # ── Cleanup khi app tắt ──
        if self._browser:
            try:
                self._browser.close()
            except Exception:
                pass
        if self._playwright:
            try:
                self._playwright.stop()
            except Exception:
                pass

# ============================================================
# GUI
# ============================================================
class App:
    def __init__(self, root):
        self.root = root; self.root.title('📦 TTS Bill — In & Tính Bill TikTok Seller')
        self.root.geometry('680x720'); self.root.minsize(520, 560)
        self.root.configure(bg='#f0f2f5')

        # Hiển thị tên file gọn gàng, không hiện đường dẫn _MEI* dài dòng
        self._cookie_real = str(DEFAULT_COOKIE) if DEFAULT_COOKIE.exists() else ''
        self._master_real = str(MASTER_DEFAULT) if MASTER_DEFAULT.exists() else ''
        # Retail: ưu tiên file cạnh exe (nếu có), fallback file nhúng
        local_retail = BASE_DIR / 'sp bán lẻ.xlsx'
        if local_retail.exists():
            self._retail_real = str(local_retail)
        elif RETAIL_DEFAULT.exists():
            self._retail_real = str(RETAIL_DEFAULT)
        else:
            self._retail_real = ''
        if getattr(sys, 'frozen', False) and DEFAULT_COOKIE.exists():
            self.cookie_path = tk.StringVar(value=f'[{DEFAULT_COOKIE.name} — đã tích hợp sẵn]')
        else:
            self.cookie_path = tk.StringVar(value=self._cookie_real)
        if getattr(sys, 'frozen', False) and MASTER_DEFAULT.exists():
            self.master_path = tk.StringVar(value=f'[{MASTER_DEFAULT.name} — đã tích hợp sẵn]')
        else:
            self.master_path = tk.StringVar(value=self._master_real)
        self.retail_path = tk.StringVar(value=self._retail_real if self._retail_real else '')
        if self._retail_real:
            p = Path(self._retail_real)
            if getattr(sys, 'frozen', False):
                self.retail_path.set(f'[{p.name} — đã tích hợp sẵn]')
        self.doc_type = tk.StringVar(value='a4')
        self.jt_orders = tk.IntVar(value=0)     # số đơn J&T Express
        self.ghn_orders = tk.IntVar(value=0)    # số đơn GHN
        self.vnp_orders = tk.IntVar(value=0)    # số đơn VietNam Post
        self.best_orders = tk.IntVar(value=0)   # số đơn Best Express
        self.viettel_orders = tk.IntVar(value=0)   # số đơn Viettel Post
        self.jtc_orders = tk.IntVar(value=0)    # số đơn J&T Cargo VN
        self.auto_print = tk.BooleanVar(value=False)  # tự động in PDF sau khi tải
        self.test_mode = tk.BooleanVar(value=True)   # chế độ test: chỉ chọn đơn, không in (mặc định ON)
        self.printer_name = tk.StringVar(value='')  # sẽ tự động chọn máy in đầu tiên
        self.paper_size = tk.StringVar(value='A4')
        self.duplex_mode = tk.StringVar(value='simplex')
        self.pages_per_sheet = tk.StringVar(value='1')
        self.output_dir = tk.StringVar(value=str(BASE_DIR / 'outputs'))
        self.schedule_mode = tk.StringVar(value='once')
        self.interval_hours = tk.IntVar(value=1)
        self.daily_times = tk.StringVar(value='08:00, 14:00, 20:00')
        self.running = False; self.scheduler_running = False; self.result_files = []
        self.stop_event = threading.Event()  # cờ dừng cho automation loop
        self._warned_sumatra = False  # tránh log cảnh báo SumatraPDF nhiều lần

        # Worker thread chuyên biệt cho Playwright (tránh lỗi thread-safety)
        self._worker = AutomationWorker()
        self._worker.start()

        self.scheduler = Scheduler(
            callback=self._execute_job_sync,
            log_cb=lambda m, t='': self.root.after(0, self.log, m, t),
            state_cb=lambda s, m: self.root.after(0, self._set_state, s, m))

        self._build_ui(); self._update_cookie_status(); self._update_master_status(); self._update_retail_status()
        os.makedirs(self.output_dir.get(), exist_ok=True)

        # Bắt sự kiện đóng cửa sổ → cleanup browser
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    # ═══════════════════════════════════════════════════
    # UI BUILDERS
    # ═══════════════════════════════════════════════════
    def _s(self, parent, title, builder):
        f = tk.LabelFrame(parent, text=title, bg='#f0f2f5', fg='#374151',
                          font=('Segoe UI', 10, 'bold'), padx=10, pady=6, bd=0, highlightthickness=1, highlightbackground='#e5e7eb')
        f.pack(fill='x', padx=0, pady=(0,3)); builder(f)

    def _row(self, parent, label, var, cmd, status_lbl_key):
        r = tk.Frame(parent, bg='#f0f2f5'); r.pack(fill='x')
        tk.Label(r, text=label, font=('Segoe UI',9), bg='#f0f2f5', fg='#6b7280', width=10, anchor='w').pack(side='left')
        tk.Entry(r, textvariable=var, font=('Segoe UI',9), relief='solid', bd=1).pack(side='left', fill='x', expand=True)
        tk.Button(r, text='📂', command=cmd, font=('Segoe UI',9), relief='flat', bg='#e5e7eb',
                  cursor='hand2', width=3).pack(side='left', padx=(4,0))
        lbl = tk.Label(parent, text='', font=('Segoe UI',8), bg='#f0f2f5', fg='#059669')
        lbl.pack(anchor='w', pady=(2,0))
        setattr(self, status_lbl_key, lbl)

    def _build_ui(self):
        # Header
        h = tk.Frame(self.root, bg='#2563eb', height=48)
        h.pack(fill='x'); h.pack_propagate(False)
        tk.Label(h, text='📦 TTS Bill', font=('Segoe UI',14,'bold'), bg='#2563eb', fg='white').pack(side='left', padx=14, pady=10)
        tk.Label(h, text='In & Tính Bill TikTok Seller', font=('Segoe UI',8), bg='#2563eb', fg='#bfdbfe').pack(side='left', pady=12)

        # ── Scrollable settings area ──
        canvas = tk.Canvas(self.root, bg='#f0f2f5', highlightthickness=0, height=340)
        scrollbar = tk.Scrollbar(self.root, orient='vertical', command=canvas.yview)
        self._settings_frame = tk.Frame(canvas, bg='#f0f2f5')
        self._settings_frame.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.create_window((0, 0), window=self._settings_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(fill='x', padx=(14, 0), pady=(4, 0))
        scrollbar.pack(side='right', fill='y', padx=(0, 14), pady=(4, 0))
        # Bind mousewheel to scroll
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), 'units')
        canvas.bind('<Enter>', lambda e: canvas.bind_all('<MouseWheel>', _on_mousewheel))
        canvas.bind('<Leave>', lambda e: canvas.unbind_all('<MouseWheel>'))

        # Build sections inside scrollable frame
        self._s(self._settings_frame, 'Cấu hình', self._build_config)
        self._s(self._settings_frame, '⚙ Tùy chọn in', self._build_options)
        self._s(self._settings_frame, '🕐 Hẹn giờ', self._build_schedule)

        # Buttons
        btn = tk.Frame(self._settings_frame, bg='#f0f2f5'); btn.pack(fill='x', padx=0, pady=(6, 4))
        self.start_btn = tk.Button(btn, text='▶ CHẠY NGAY', command=self._start_once,
            font=('Segoe UI',10,'bold'), bg='#2563eb', fg='white', relief='flat', cursor='hand2', padx=12, pady=6,
            activebackground='#1d4ed8', activeforeground='white')
        self.start_btn.pack(side='left', fill='x', expand=True, padx=(0, 2))
        self.sched_btn = tk.Button(btn, text='⏰ CHẠY LỊCH', command=self._start_schedule,
            font=('Segoe UI',10,'bold'), bg='#059669', fg='white', relief='flat', cursor='hand2', padx=12, pady=6,
            activebackground='#047857', activeforeground='white')
        self.sched_btn.pack(side='left', fill='x', expand=True, padx=(2, 2))
        self.stop_btn = tk.Button(btn, text='⏹ DỪNG', command=self._stop_all,
            font=('Segoe UI',10,'bold'), bg='#e5e7eb', fg='#6b7280', relief='flat', cursor='hand2', padx=12, pady=6,
            state='disabled', activebackground='#d1d5db')
        self.stop_btn.pack(side='left', fill='x', expand=True, padx=(2, 0))

        # Status bar
        sf = tk.Frame(self.root, bg='#f0f2f5'); sf.pack(fill='x', padx=14, pady=(4, 0))
        self.status_bar = tk.Label(sf, text='🟢 Sẵn sàng', font=('Segoe UI',9,'bold'),
                                   bg='#f0f2f5', fg='#059669', anchor='w')
        self.status_bar.pack(fill='x')

        # LOG — fills remaining space
        self._build_log()

        # Results — compact at bottom
        self._build_results()

    def _build_config(self, parent):
        self._row(parent, 'Cookie', self.cookie_path, self._select_cookie, 'cookie_status')
        self._row(parent, 'Master (Combo)', self.master_path, self._select_master, 'master_status')
        self._row(parent, 'Bán lẻ', self.retail_path, self._select_retail, 'retail_status')
        self._row(parent, 'Thư mục lưu', self.output_dir, self._select_output_dir, 'output_status')
        self.output_status.config(text='✅ Sẵn sàng', fg='#059669')

    def _build_options(self, parent):
        # Dòng 1: Số đơn J&T
        r1 = tk.Frame(parent, bg='#f0f2f5'); r1.pack(fill='x')
        tk.Label(r1, text='J&T Express:', font=('Segoe UI',9), bg='#f0f2f5', fg='#6b7280', width=12, anchor='w').pack(side='left')
        tk.Spinbox(r1, from_=0, to=9999, width=6, textvariable=self.jt_orders, font=('Segoe UI',9)).pack(side='left', padx=(0,0))
        tk.Label(r1, text='đơn  (0 = bỏ qua)', font=('Segoe UI',8), bg='#f0f2f5', fg='#9ca3af').pack(side='left', padx=6)
        # Dòng 2: Số đơn GHN
        r2 = tk.Frame(parent, bg='#f0f2f5'); r2.pack(fill='x', pady=(2,0))
        tk.Label(r2, text='GHN:', font=('Segoe UI',9), bg='#f0f2f5', fg='#6b7280', width=12, anchor='w').pack(side='left')
        tk.Spinbox(r2, from_=0, to=9999, width=6, textvariable=self.ghn_orders, font=('Segoe UI',9)).pack(side='left', padx=(0,0))
        tk.Label(r2, text='đơn  (0 = bỏ qua)', font=('Segoe UI',8), bg='#f0f2f5', fg='#9ca3af').pack(side='left', padx=6)
        # Dòng 3: Số đơn VietNam Post
        r3 = tk.Frame(parent, bg='#f0f2f5'); r3.pack(fill='x', pady=(2,0))
        tk.Label(r3, text='VietNam Post:', font=('Segoe UI',9), bg='#f0f2f5', fg='#6b7280', width=12, anchor='w').pack(side='left')
        tk.Spinbox(r3, from_=0, to=9999, width=6, textvariable=self.vnp_orders, font=('Segoe UI',9)).pack(side='left', padx=(0,0))
        tk.Label(r3, text='đơn  (0 = bỏ qua)', font=('Segoe UI',8), bg='#f0f2f5', fg='#9ca3af').pack(side='left', padx=6)
        # Dòng 4: Số đơn Best Express
        r4 = tk.Frame(parent, bg='#f0f2f5'); r4.pack(fill='x', pady=(2,0))
        tk.Label(r4, text='Best Express:', font=('Segoe UI',9), bg='#f0f2f5', fg='#6b7280', width=12, anchor='w').pack(side='left')
        tk.Spinbox(r4, from_=0, to=9999, width=6, textvariable=self.best_orders, font=('Segoe UI',9)).pack(side='left', padx=(0,0))
        tk.Label(r4, text='đơn  (0 = bỏ qua)', font=('Segoe UI',8), bg='#f0f2f5', fg='#9ca3af').pack(side='left', padx=6)
        # Dòng 5: Số đơn Viettel Post
        r5 = tk.Frame(parent, bg='#f0f2f5'); r5.pack(fill='x', pady=(2,0))
        tk.Label(r5, text='Viettel Post:', font=('Segoe UI',9), bg='#f0f2f5', fg='#6b7280', width=12, anchor='w').pack(side='left')
        tk.Spinbox(r5, from_=0, to=9999, width=6, textvariable=self.viettel_orders, font=('Segoe UI',9)).pack(side='left', padx=(0,0))
        tk.Label(r5, text='đơn  (0 = bỏ qua)', font=('Segoe UI',8), bg='#f0f2f5', fg='#9ca3af').pack(side='left', padx=6)
        # Dòng 6: Số đơn J&T Cargo VN
        r6 = tk.Frame(parent, bg='#f0f2f5'); r6.pack(fill='x', pady=(2,0))
        tk.Label(r6, text='J&T Cargo VN:', font=('Segoe UI',9), bg='#f0f2f5', fg='#6b7280', width=12, anchor='w').pack(side='left')
        tk.Spinbox(r6, from_=0, to=9999, width=6, textvariable=self.jtc_orders, font=('Segoe UI',9)).pack(side='left', padx=(0,0))
        tk.Label(r6, text='đơn  (0 = bỏ qua)', font=('Segoe UI',8), bg='#f0f2f5', fg='#9ca3af').pack(side='left', padx=6)
        tk.Label(parent, text='📑 Luôn in: Danh sách lấy hàng (A4) + Đơn vận chuyển (A6)',
                 font=('Segoe UI',8), bg='#f0f2f5', fg='#6b7280').pack(anchor='w', pady=(4,0))
        # In tự động
        r3 = tk.Frame(parent, bg='#f0f2f5'); r3.pack(fill='x', pady=(4,0))
        tk.Checkbutton(r3, text='🖨️ In tự động ra máy in', variable=self.auto_print,
                       bg='#f0f2f5', font=('Segoe UI',9), activebackground='#f0f2f5').pack(side='left')
        # Combobox chọn máy in (tự động lấy danh sách)
        printers = self._get_printers()
        self.printer_combo = ttk.Combobox(r3, textvariable=self.printer_name, values=printers,
                                          font=('Segoe UI',8), width=35, state='readonly')
        self.printer_combo.pack(side='left', padx=(6, 0))
        if printers:
            self.printer_combo.current(0)  # chọn máy in đầu tiên
        # Nút refresh danh sách máy in
        tk.Button(r3, text='🔄', command=self._refresh_printers, font=('Segoe UI',7),
                  relief='flat', bg='#e5e7eb', cursor='hand2', width=3).pack(side='left', padx=(4, 0))
        r4 = tk.Frame(parent, bg='#f0f2f5'); r4.pack(fill='x', pady=(2, 0))
        tk.Label(r4, text='   Khổ giấy:', font=('Segoe UI',8), bg='#f0f2f5', fg='#6b7280').pack(side='left')
        self.paper_combo = ttk.Combobox(r4, textvariable=self.paper_size,
                                         values=['A4', 'A3', 'Letter', 'Legal'],
                                         font=('Segoe UI',8), width=8, state='readonly')
        self.paper_combo.pack(side='left', padx=(2, 6))
        tk.Label(r4, text='In mặt:', font=('Segoe UI',8), bg='#f0f2f5', fg='#6b7280').pack(side='left')
        self.duplex_combo = ttk.Combobox(r4, textvariable=self.duplex_mode,
                                          values=['simplex', 'longedge', 'shortedge'],
                                          font=('Segoe UI',8), width=10, state='readonly')
        self.duplex_combo.pack(side='left', padx=(2, 6))
        tk.Label(r4, text='Trang/tờ:', font=('Segoe UI',8), bg='#f0f2f5', fg='#6b7280').pack(side='left')
        self.pages_combo = ttk.Combobox(r4, textvariable=self.pages_per_sheet,
                                         values=['1', '2', '4', '6', '9'],
                                         font=('Segoe UI',8), width=4, state='readonly')
        self.pages_combo.pack(side='left', padx=(2, 6))
        # Bind: khi đổi máy in → tự động cập nhật khổ giấy hỗ trợ
        self.printer_combo.bind('<<ComboboxSelected>>', lambda e: self._on_printer_changed())
        # Test mode
        r_test = tk.Frame(parent, bg='#f0f2f5'); r_test.pack(fill='x', pady=(2,0))
        tk.Checkbutton(r_test, text='🧪 Test mode (chỉ chọn đơn, không in)', variable=self.test_mode,
                       bg='#f0f2f5', font=('Segoe UI',9), fg='#d97706', activebackground='#f0f2f5').pack(side='left')

    def _build_schedule(self, parent):
        modes = [('Chạy 1 lần','once'), ('Lặp mỗi N giờ','interval'), ('Giờ cố định trong ngày','daily')]
        for i, (lbl, mode) in enumerate(modes):
            tk.Radiobutton(parent, text=lbl, variable=self.schedule_mode, value=mode,
                           bg='#f0f2f5', font=('Segoe UI',9), command=self._on_sched_change,
                           activebackground='#f0f2f5').grid(row=i, column=0, sticky='w', pady=1)

        self.intv_f = tk.Frame(parent, bg='#f0f2f5')
        tk.Label(self.intv_f, text='Mỗi', bg='#f0f2f5', font=('Segoe UI',9)).pack(side='left')
        tk.Spinbox(self.intv_f, from_=1, to=24, width=3, textvariable=self.interval_hours, font=('Segoe UI',9)).pack(side='left', padx=4)
        tk.Label(self.intv_f, text='giờ', bg='#f0f2f5', font=('Segoe UI',9)).pack(side='left')

        self.day_f = tk.Frame(parent, bg='#f0f2f5')
        tk.Label(self.day_f, text='Lúc:', bg='#f0f2f5', font=('Segoe UI',9)).pack(side='left')
        tk.Entry(self.day_f, textvariable=self.daily_times, width=22, font=('Segoe UI',9)).pack(side='left', padx=4)
        tk.Label(self.day_f, text='VD: 08:00, 14:00', bg='#f0f2f5', font=('Segoe UI',8), fg='#9ca3af').pack(side='left')
        self._on_sched_change()

    def _build_log(self):
        lf = tk.LabelFrame(self.root, text='📋 Log', bg='#f0f2f5', fg='#374151',
                           font=('Segoe UI',10,'bold'), padx=4, pady=4, bd=0, highlightthickness=1, highlightbackground='#e5e7eb')
        lf.pack(fill='both', expand=True, padx=14, pady=(6,0))

        # Toolbar nhỏ: clear log
        tb = tk.Frame(lf, bg='#f0f2f5'); tb.pack(fill='x', pady=(0,2))
        tk.Button(tb, text='Xóa log', command=self._clear_log, font=('Segoe UI',8),
                  relief='flat', bg='#e5e7eb', cursor='hand2', padx=8).pack(side='right')

        self.log_text = tk.Text(lf, font=('Consolas',9),
                                bg='#1a1d23', fg='#d4d4d4', relief='flat',
                                padx=12, pady=10, state='disabled', wrap='word',
                                selectbackground='#264f78', selectforeground='white')
        self.log_text.pack(fill='both', expand=True)

        # Tags màu
        for tag, cfg in {
            'ts':     '#6b7280',
            'ok':     '#4ade80',
            'err':    '#f87171',
            'warn':   '#fbbf24',
            'info':   '#60a5fa',
            'batch':  '#c084fc',
            'result': '#f472b6',
            'dim':    '#6b7280',
        }.items():
            self.log_text.tag_config(tag, foreground=cfg)
        self.log_text.tag_config('bold_ok', foreground='#4ade80', font=('Consolas',9,'bold'))

    def _build_results(self):
        rf = tk.LabelFrame(self.root, text='📁 Kết quả — double-click để mở file', bg='#f0f2f5', fg='#374151',
                           font=('Segoe UI',10,'bold'), padx=6, pady=6, bd=0, highlightthickness=1, highlightbackground='#e5e7eb')
        rf.pack(fill='x', padx=14, pady=(6,10))

        self.result_list = tk.Listbox(rf, height=4, font=('Segoe UI',9), relief='flat',
                                      selectbackground='#2563eb', selectforeground='white',
                                      activestyle='none')
        self.result_list.pack(fill='x'); self.result_list.bind('<Double-Button-1>', self._open_result)

        br = tk.Frame(rf, bg='#f0f2f5'); br.pack(fill='x', pady=(4,0))
        tk.Button(br, text='📂 Mở thư mục', command=self._open_output_dir,
                  font=('Segoe UI',9), relief='flat', bg='#e5e7eb', cursor='hand2').pack(side='right')
        tk.Button(br, text='🗑 Xóa hết', command=self._clear_results,
                  font=('Segoe UI',9), relief='flat', bg='#e5e7eb', cursor='hand2').pack(side='right', padx=4)

    # ═══════════════════════════════════════════════════
    # ACTIONS
    # ═══════════════════════════════════════════════════
    def _on_sched_change(self):
        self.intv_f.grid_forget(); self.day_f.grid_forget()
        if self.schedule_mode.get() == 'interval': self.intv_f.grid(row=1, column=0, sticky='w', padx=(20,0), pady=(4,0))
        elif self.schedule_mode.get() == 'daily': self.day_f.grid(row=2, column=0, sticky='w', padx=(20,0), pady=(4,0))

    def _select_cookie(self):
        p = filedialog.askopenfilename(filetypes=[('JSON','*.json')])
        if p: self._cookie_real = p; self.cookie_path.set(p); self._update_cookie_status()
    def _select_master(self):
        p = filedialog.askopenfilename(filetypes=[('Excel','*.xlsx')])
        if p: self._master_real = p; self.master_path.set(p); self._update_master_status()
    def _select_retail(self):
        p = filedialog.askopenfilename(filetypes=[('Excel','*.xlsx')])
        if p: self._retail_real = p; self.retail_path.set(p); self._update_retail_status()
    def _select_output_dir(self):
        p = filedialog.askdirectory()
        if p: self.output_dir.set(p)

    def _update_cookie_status(self):
        p = Path(self._cookie_real) if self._cookie_real else Path('')
        if self._cookie_real and p.exists():
            try:
                d = json.loads(p.read_text(encoding='utf-8'))
                cs = d.get('cookies', d); n = len(cs) if isinstance(cs, list) else 0
                self.cookie_status.config(text=f'✅ {n} cookies', fg='#059669')
            except: self.cookie_status.config(text='⚠ File lỗi', fg='#d97706')
        else: self.cookie_status.config(text='❌ Chưa chọn', fg='#dc2626')

    def _update_master_status(self):
        p = Path(self._master_real) if self._master_real else Path('')
        if self._master_real and p.exists() and p.suffix.lower() == '.xlsx':
            try:
                from openpyxl import load_workbook
                wb = load_workbook(str(p), data_only=True); n = wb.active.max_row - 1; wb.close()
                self.master_status.config(text=f'✅ {n} SKU', fg='#059669')
            except: self.master_status.config(text='⚠ File lỗi', fg='#d97706')
        else: self.master_status.config(text='❌ Chưa chọn', fg='#dc2626')

    def _update_retail_status(self):
        p = Path(self._retail_real) if self._retail_real else Path('')
        if self._retail_real and p.exists():
            try:
                from openpyxl import load_workbook
                wb = load_workbook(str(p), data_only=True); n = wb.active.max_row - 1; wb.close()
                self.retail_status.config(text=f'✅ {n} SKU', fg='#059669')
            except: self.retail_status.config(text='⚠ File lỗi', fg='#d97706')
        else: self.retail_status.config(text='❌ Chưa chọn', fg='#dc2626')

    def log(self, msg, tag=''):
        self.log_text.config(state='normal')
        ts = datetime.now().strftime('%H:%M:%S')
        self.log_text.insert('end', ts + '  ', 'ts')
        self.log_text.insert('end', msg + '\n', tag)
        self.log_text.see('end'); self.log_text.config(state='disabled')

    def _set_state(self, step, msg):
        c = {'idle':'#059669','running':'#d97706','waiting':'#2563eb','done':'#059669','error':'#dc2626'}
        self.status_bar.config(text=msg, fg=c.get(step,'#333'))

    def _set_buttons(self, s):
        if s in ('running','scheduled'):
            self.start_btn.config(state='disabled', bg='#93c5fd', text='▶ ĐANG CHẠY')
            self.sched_btn.config(state='disabled', bg='#6ee7b7')
            self.stop_btn.config(state='normal', bg='#fca5a5', fg='#7f1d1d')
        else:
            self.start_btn.config(state='normal', bg='#2563eb', text='▶ CHẠY NGAY')
            self.sched_btn.config(state='normal', bg='#059669', text='⏰ CHẠY LỊCH')
            self.stop_btn.config(state='disabled', bg='#e5e7eb', fg='#6b7280', text='⏹ DỪNG')

    def _start_once(self):
        if self.running: return
        if not Path(self._cookie_real).exists(): messagebox.showerror('Lỗi','Chọn file cookie JSON hợp lệ.'); return
        self.running = True; self._set_buttons('running'); self._clear_results(); self._clear_log()
        self.log('▶ Bắt đầu...', 'bold_ok')
        self._worker.run_job(self._execute_job)

    def _start_schedule(self):
        if not Path(self._cookie_real).exists(): messagebox.showerror('Lỗi','Chọn file cookie JSON hợp lệ.'); return
        mode = self.schedule_mode.get()
        if mode == 'once': self._start_once(); return
        if mode == 'daily':
            times = [t.strip() for t in self.daily_times.get().split(',') if t.strip()]
            if not times: messagebox.showerror('Lỗi','Nhập ít nhất 1 giờ (VD: 08:00).'); return
        self.scheduler.configure(mode=mode, interval_hours=self.interval_hours.get(),
            daily_times=[t.strip() for t in self.daily_times.get().split(',') if t.strip()])
        self._clear_log(); self._clear_results()
        self.log(f'⏰ Hẹn giờ: {mode}', 'info')
        if mode == 'interval': self.log(f'   Chạy mỗi {self.interval_hours.get()} giờ', 'info')
        elif mode == 'daily': self.log(f'   Chạy lúc {self.daily_times.get()}', 'info')
        self.scheduler_running = True; self._set_buttons('scheduled'); self.scheduler.start()

    def _stop_all(self):
        self.scheduler.stop(); self.scheduler_running = False; self.running = False
        self.stop_event.set()  # báo hiệu dừng cho automation loop
        self._worker.stop_job()  # dừng job đang chạy (giữ browser)
        self._set_buttons('idle'); self._set_state('idle','⏹ Đã dừng'); self.log('⏹ Đã dừng', 'warn')

    def _execute_job(self, playwright=None, browser=None, job_stop=None):
        """Thực thi 1 job automation + calculator. Được gọi từ AutomationWorker thread."""
        try:
            cookie = self._cookie_real; base_dir = self.output_dir.get()
            master = self._master_real; retail = self._retail_real
            doc = self.doc_type.get()
            # Tự động tạo thư mục con theo ngày để gọn gàng từng ngày
            today_str = datetime.now().strftime('%Y-%m-%d')
            out_dir = str(Path(base_dir) / today_str)
            os.makedirs(out_dir, exist_ok=True)

            # Xác định danh sách carrier cần xử lý
            all_carriers = [
                ('J&T',           self.jt_orders.get()),
                ('GHN',           self.ghn_orders.get()),
                ('VietNam Post',  self.vnp_orders.get()),
                ('Best Express',  self.best_orders.get()),
                ('Viettel Post',  self.viettel_orders.get()),
                ('J&T Cargo VN',  self.jtc_orders.get()),
            ]

            # Nếu tất cả đều = 0 → chạy tất cả carrier, mỗi carrier in tất cả đơn
            if all(v == 0 for _, v in all_carriers):
                carriers_to_process = [(name, 0) for name, _ in all_carriers]
            else:
                carriers_to_process = [(name, v) for name, v in all_carriers if v != 0]

            all_pdf_paths = []
            all_results = []  # gom kết quả calculator từ tất cả carrier
            stop_signal = job_stop if job_stop is not None else self.stop_event

            for carrier, count in carriers_to_process:
                self.stop_event.clear()
                carrier_display = carrier if carrier else 'tất cả'
                count_display = f'{count}' if count > 0 else 'tất cả'
                self.log(f'📥 Tải PDF [{carrier_display}] ({count_display} đơn)...', 'info')

                pdf_paths, pw, br = run_automation(cookie, doc, out_dir, count,
                    lambda m, t='': self.root.after(0, self.log, m, t),
                    lambda s, m: self.root.after(0, self._set_state, s, m),
                    stop_signal,
                    existing_playwright=playwright, existing_browser=browser,
                    carrier=carrier, test_mode=self.test_mode.get())

                # Giữ browser cho carrier tiếp theo
                if pw: playwright = pw
                if br: browser = br

                for p in pdf_paths:
                    if Path(p).exists(): self._add_result(f'📄 {Path(p).name}')
                all_pdf_paths.extend(pdf_paths)
                self.log(f'📥 [{carrier_display}]: đã tải {len(pdf_paths)} file PDF', 'ok' if pdf_paths else 'warn')

                # Calculator riêng cho từng carrier
                if pdf_paths:
                    self.log(f'📊 Đang tính bill [{carrier_display}]...', 'info')
                    carrier_results = run_calculator(pdf_paths, out_dir, master, retail,
                        lambda m, t='': self.root.after(0, self.log, m, t),
                        carrier=carrier)
                    for r in carrier_results:
                        for key, lbl in [('excel','📊'),('pdf','📕'),('pdf_grouped','📋')]:
                            fp = r['files'].get(key)
                            if fp and Path(fp).exists(): self._add_result(f'{lbl} {Path(fp).name}')
                    all_results.extend(carrier_results)

            # Step 3: In tự động ra máy in (nếu bật)
            if self.auto_print.get():
                printer = self.printer_name.get().strip()
                self.log(f'🖨️ Đang in ra "{printer}"...', 'info')

                # Gom file theo timestamp để in xen kẽ: shipping label → Excel
                # Timestamp format trong tên file: MM-DD_HH-MM-SS
                import re as _re
                groups = {}  # {ts: {'shipping': [paths], 'excel': [paths]}}

                for p in all_pdf_paths:
                    name = Path(p).name
                    if not (Path(p).exists() and ('shipping' in name.lower() or 'vận chuyển' in name.lower())):
                        continue
                    m = _re.search(r'(\d{2}-\d{2}_\d{2}-\d{2}-\d{2})', name)
                    ts = m.group(1) if m else '0'
                    groups.setdefault(ts, {'shipping': [], 'excel': []})['shipping'].append(str(p))

                for r in all_results:
                    excel_path = r['files'].get('excel')
                    if not (excel_path and Path(excel_path).exists()):
                        continue
                    name = Path(excel_path).name
                    m = _re.search(r'(\d{2}-\d{2}_\d{2}-\d{2}-\d{2})', name)
                    ts = m.group(1) if m else '0'
                    groups.setdefault(ts, {'shipping': [], 'excel': []})['excel'].append(excel_path)

                # In theo thứ tự: shipping label 1 → Excel 1 → shipping label 2 → Excel 2 ...
                label_settings = f'paper={self.paper_size.get()},duplex=simplex'
                printed = 0
                for ts in sorted(groups.keys()):
                    g = groups[ts]
                    # In shipping label trước
                    for fp in g['shipping']:
                        try:
                            self._print_file(fp, printer, pdf_settings=label_settings)
                            self.log(f'  ✓ Đã gửi in: {Path(fp).name}', 'ok')
                            printed += 1
                        except Exception as e:
                            self.log(f'  ✗ Lỗi in {Path(fp).name}: {e}', 'err')
                    # In Excel báo cáo gộp SKU ngay sau shipping label cùng batch
                    for fp in g['excel']:
                        try:
                            self._print_file(fp, printer, pdf_settings='')
                            self.log(f'  ✓ Đã gửi in: {Path(fp).name}', 'ok')
                            printed += 1
                        except Exception as e:
                            self.log(f'  ✗ Lỗi in {Path(fp).name}: {e}', 'err')

                if printed == 0:
                    self.log('  ⚠ Không có file nào để in', 'warn')

            self.log('🏁 HOÀN THÀNH!', 'bold_ok')
            self._set_state('done', f'✅ Hoàn thành lúc {datetime.now().strftime("%H:%M:%S")}')
            # Trả về browser/playwright để worker lưu lại cho lần sau
            return {'playwright': pw, 'browser': br, 'pdf_paths': all_pdf_paths}
        except Exception as e:
            self.log(f'✗ Lỗi: {e}', 'err'); self._set_state('error', f'✗ {e}')
            return {}
        finally:
            self.running = False
            if self.schedule_mode.get() == 'once' or not self.scheduler_running:
                self.root.after(0, lambda: self._set_buttons('idle'))

    def _execute_job_sync(self):
        """Wrapper cho scheduler — gửi job vào worker và đợi hoàn thành."""
        done = threading.Event()
        def job_wrapper(pw, br, stop):
            try:
                return self._execute_job(pw, br, stop)
            finally:
                self.root.after(0, done.set)
        self._worker.run_job(job_wrapper)
        done.wait()  # block cho đến khi job hoàn thành

    def _add_result(self, text):
        self.root.after(0, lambda: self.result_list.insert('end', text)); self.result_files.append(text)
    def _clear_results(self): self.result_list.delete(0,'end'); self.result_files.clear()
    def _clear_log(self):
        self.log_text.config(state='normal'); self.log_text.delete('1.0','end'); self.log_text.config(state='disabled')

    def _open_result(self, event):
        sel = self.result_list.curselection()
        if sel:
            item = self.result_list.get(sel[0])
            for pfx in ['📄 ','📊 ','📕 ','📋 ']:
                if item.startswith(pfx): item = item[len(pfx):]; break
            fp = str(Path(self.output_dir.get()) / item)
            if Path(fp).exists(): os.startfile(fp)

    def _open_output_dir(self):
        base = self.output_dir.get()
        d = Path(base) / datetime.now().strftime('%Y-%m-%d')
        if d.exists(): os.startfile(str(d))
        elif Path(base).exists(): os.startfile(base)

    def _print_file(self, file_path, printer_name, pdf_settings='paper=A4'):
        """In file (PDF hoặc Excel) ra máy in chỉ định (Windows).
        pdf_settings: chuỗi cài đặt (vd: 'paper=A4,pagespersheet=2,duplex=simplex')"""
        import subprocess, os as _os
        fp = str(file_path)

        # ── PDF: tự merge 2-up nếu cần, rồi in qua SumatraPDF ──
        if fp.lower().endswith('.pdf'):
            sumatra_exe = None
            sumatra_paths = [
                _os.path.join(_os.path.dirname(sys.executable), 'SumatraPDF.exe') if getattr(sys, 'frozen', False) else r'C:\Users\thanh\AppData\Local\SumatraPDF\SumatraPDF.exe',
                r'C:\Users\thanh\AppData\Local\SumatraPDF\SumatraPDF.exe',
                r'C:\Program Files\SumatraPDF\SumatraPDF.exe',
                r'C:\Program Files (x86)\SumatraPDF\SumatraPDF.exe',
            ]
            for sp in sumatra_paths:
                if Path(sp).exists():
                    sumatra_exe = sp
                    break

            # Luôn tự merge 2-up cho PDF (nếu có ≥2 trang) — layout Letter Portrait 130%
            print_path = fp
            temp_merged = None
            try:
                temp_merged = self._merge_pdf_2up(fp)
                if temp_merged:
                    print_path = temp_merged
            except Exception:
                pass  # fallback: in thẳng file gốc

            if sumatra_exe:
                cmd = [sumatra_exe, '-print-to', printer_name, print_path]
                if pdf_settings:
                    cmd += ['-print-settings', pdf_settings]
                subprocess.run(cmd, check=False, timeout=60)
            else:
                self._warn_sumatra()
                subprocess.run(['powershell', '-Command',
                    f'Start-Process -FilePath "{print_path}" -Verb Print'],
                    capture_output=True, timeout=30)

            # Dọn file tạm
            if temp_merged:
                def _cleanup(p=temp_merged):
                    import time; time.sleep(5)
                    try: _os.remove(p)
                    except: pass
                threading.Timer(5, _cleanup).start()
            return

        # ── Excel: in thẳng qua COM automation (A4 mặc định) ──
        if fp.lower().endswith('.xlsx') or fp.lower().endswith('.xls'):
            try:
                import pythoncom, win32com.client
                pythoncom.CoInitialize()
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                workbook = excel.Workbooks.Open(_os.path.abspath(fp))
                workbook.PrintOut(ActivePrinter=printer_name)
                workbook.Close(False)
                excel.Quit()
                pythoncom.CoUninitialize()
                return
            except Exception:
                try: pythoncom.CoUninitialize()
                except: pass

        # ── Fallback: PowerShell ──
        subprocess.run(['powershell', '-Command',
            f'Start-Process -FilePath "{fp}" -Verb Print'],
            capture_output=True, timeout=30)

    def _merge_pdf_2up(self, pdf_path):
        """Gộp PDF thành các tờ A4 Landscape (2 label nằm ngang cạnh nhau/trang, scale 70%).
        Mỗi cặp 2 trang PDF → 1 tờ A4. Trả về đường dẫn file đã merge, hoặc None."""
        import os as _os
        try:
            from pypdf import PdfReader, PdfWriter, PageObject, Transformation
        except ImportError:
            return None
        try:
            reader = PdfReader(pdf_path)
            if len(reader.pages) < 2:
                return None

            # A4 Landscape: 842 x 595 pts
            canvas_w, canvas_h = 842, 595
            margin_left = 14   # 5mm
            margin_top = 28    # 10mm
            gap = -14          # dồn label phải sát trái
            scale_factor = 0.70

            writer = PdfWriter()
            avail_w = canvas_w - 2*margin_left - gap
            half_w = avail_w / 2

            # Xử lý từng cặp 2 trang
            for pair_start in range(0, len(reader.pages), 2):
                pair = reader.pages[pair_start:pair_start+2]
                is_last_odd = len(pair) == 1  # label lẻ cuối cùng
                canvas = PageObject.create_blank_page(width=canvas_w, height=canvas_h)
                for i, page in enumerate(pair):
                    pw = float(page.mediabox.width)
                    ph = float(page.mediabox.height)
                    scale = min(half_w / pw, (canvas_h - 2*margin_top) / ph) * scale_factor
                    sw, sh = pw * scale, ph * scale
                    tx = margin_left + i * (half_w + gap)
                    ty = canvas_h - margin_top - sh
                    canvas.merge_transformed_page(page,
                        Transformation().scale(scale).translate(tx / scale, ty / scale))
                writer.add_page(canvas)

            temp_path = pdf_path + '.2up.pdf'
            with open(temp_path, 'wb') as f:
                writer.write(f)
            return temp_path
        except Exception:
            return None

    def _get_printers(self) -> list[str]:
        """Trả về danh sách máy in đang có trên hệ thống (Windows)."""
        printers = []
        # Cách 1: win32print (nếu có)
        try:
            import win32print
            for info in win32print.EnumPrinters(win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS, None, 1):
                name = info[2].strip() if info[2] else ''
                if name:
                    printers.append(name)
        except Exception:
            pass

        # Cách 2: PowerShell fallback
        if not printers:
            try:
                import subprocess
                result = subprocess.run(
                    ['powershell', '-Command',
                     "Get-Printer | Select-Object -ExpandProperty Name | Where-Object { $_ -notlike '*Microsoft*' -and $_ -notlike '*Fax*' -and $_ -notlike '*OneNote*' -and $_ -notlike '*XPS*' }"],
                    capture_output=True, text=True, timeout=10)
                for line in result.stdout.strip().split('\n'):
                    name = line.strip()
                    if name and name not in printers:
                        printers.append(name)
            except Exception:
                pass

        return printers

    def _refresh_printers(self):
        """Làm mới danh sách máy in trong combobox."""
        printers = self._get_printers()
        self.printer_combo['values'] = printers
        if printers:
            self.printer_combo.current(0)
            self.printer_name.set(printers[0])
        self._on_printer_changed()

    def _on_printer_changed(self):
        """Khi đổi máy in → cập nhật danh sách khổ giấy được hỗ trợ."""
        printer = self.printer_name.get()
        if not printer:
            return
        papers = self._get_printer_papers(printer)
        if papers:
            self.paper_combo['values'] = papers
            # Giữ A4 nếu có, nếu không thì chọn cái đầu
            if 'A4' in papers:
                self.paper_size.set('A4')
            else:
                self.paper_size.set(papers[0])

    def _get_printer_papers(self, printer_name: str) -> list[str]:
        """Lấy danh sách khổ giấy được hỗ trợ bởi máy in (Windows)."""
        papers = []
        # Dùng win32print
        try:
            import win32print, win32con
            hprinter = win32print.OpenPrinter(printer_name)
            try:
                # DC_PAPERS = 2, DC_PAPERNAMES = 16
                paper_ids = win32print.DeviceCapabilities(printer_name, None, 2)
                paper_names = win32print.DeviceCapabilities(printer_name, None, 16)
                if paper_names and paper_ids:
                    # paper_names là tuple các tên, paper_ids là tuple các ID
                    seen = set()
                    for i, pid in enumerate(paper_ids):
                        if i < len(paper_names) and paper_names[i]:
                            name = paper_names[i].strip()
                            if name and name not in seen:
                                papers.append(name)
                                seen.add(name)
            finally:
                win32print.ClosePrinter(hprinter)
        except Exception:
            pass

        # Fallback: PowerShell
        if not papers:
            try:
                import subprocess
                ps = f'''Get-PrinterProperty -PrinterName "{printer_name}" | Select-Object -ExpandProperty PaperSizes 2>$null'''
                result = subprocess.run(['powershell', '-Command', ps],
                                        capture_output=True, text=True, timeout=10)
                for line in result.stdout.strip().split('\n'):
                    name = line.strip()
                    if name:
                        papers.append(name)
            except Exception:
                pass

        # Fallback mặc định
        if not papers:
            papers = ['A4', 'A3', 'Letter', 'Legal']
        return papers

    def _build_pdf_settings(self) -> str:
        """Dựng chuỗi cài đặt SumatraPDF từ các combobox."""
        parts = [f'paper={self.paper_size.get()}',
                 f'duplex={self.duplex_mode.get()}']
        pages = self.pages_per_sheet.get()
        if pages != '1':
            parts.append(f'pagespersheet={pages}')
        return ','.join(parts)

    def _warn_sumatra(self):
        """Cảnh báo 1 lần nếu chưa cài SumatraPDF."""
        if not getattr(self, '_warned_sumatra', False):
            self.root.after(0, self.log,
                '⚠ Chưa cài SumatraPDF — tải tại: https://www.sumatrapdfreader.org', 'warn')
            self._warned_sumatra = True

    def _on_close(self):
        """Dọn dẹp khi tắt app: dừng scheduler, đóng browser, hủy cửa sổ."""
        self.scheduler.stop()
        self.scheduler_running = False
        self.running = False
        self.stop_event.set()
        self._worker.shutdown()  # đóng browser + dừng worker thread
        self.root.destroy()

if __name__ == '__main__':
    root = tk.Tk()
    App(root)
    root.mainloop()
