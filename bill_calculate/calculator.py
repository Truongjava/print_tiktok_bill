"""
calculator.py — Module xử lý PDF danh sách sản phẩm
==================================================
Đối chiếu Picking List PDF với master_data.xlsx để tính số lượng thực tế.
Mỗi Seller SKU trong PDF được tra cứu trong master_data, sau đó nhân số lượng.
"""
import os
import re
from collections import defaultdict
from datetime import datetime

from fpdf import FPDF
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# MASTER DATA
# ============================================================

def load_master_data(master_path: str) -> list[dict]:
    """
    Đọc file master_data.xlsx (hoặc mã combo.xlsx).
    Tự động nhận diện cột theo tên header, không phụ thuộc vị trí.
    Trả về list các dòng, mỗi dòng là dict:
      {seller_sku, sku, qty, qty_sold, promo_qty}
    """
    from openpyxl import load_workbook

    wb = load_workbook(master_path, data_only=True)
    ws = wb.active

    # ── Đọc header để xác định vị trí các cột ──
    headers = {}
    for ci, cell in enumerate(ws[1], 1):
        val = str(cell.value).lower().strip() if cell.value else ''
        headers[ci] = val

    def find_col(keywords: list[str]) -> int:
        """Tìm cột theo từ khóa (không phân biệt hoa thường)."""
        for ci, h in headers.items():
            for kw in keywords:
                if kw in h:
                    return ci
        return 0

    col_seller_sku = find_col(['combo', 'seller sku', 'seller_sku', 'mã combo', 'ma combo']) or 1
    col_sku = find_col(['sku']) or 2
    col_qty = find_col(['sl', 'qty', 'số lượng', 'so luong']) or 3
    col_qty_sold = find_col(['sl bán', 'sl ban', 'qty sold', 'qty_sold', 'bán', 'ban']) or 4
    col_promo = find_col(['sl km', 'promo qty', 'promo_qty', 'km']) or 5
    col_unit = find_col(['đơn vị tính', 'don vi tinh', 'đơn vị', 'don vi', 'dvt', 'unit']) or 0

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[col_seller_sku - 1] is None:
            continue
        seller_sku = str(row[col_seller_sku - 1]).strip()
        sku = str(row[col_sku - 1]).strip() if row[col_sku - 1] is not None else seller_sku
        qty = int(row[col_qty - 1]) if row[col_qty - 1] is not None else 0
        qty_sold = int(row[col_qty_sold - 1]) if row[col_qty_sold - 1] is not None else 0
        promo_qty = int(row[col_promo - 1]) if row[col_promo - 1] is not None else 0
        unit = str(row[col_unit - 1]).strip() if col_unit and row[col_unit - 1] is not None else ''
        rows.append({
            "seller_sku": seller_sku,
            "sku": sku,
            "qty": qty,
            "qty_sold": qty_sold,
            "promo_qty": promo_qty,
            "unit": unit,
        })
    wb.close()
    return rows


def load_retail_data(retail_path: str) -> dict[str, dict]:
    """
    Doc file san pham ban le.xlsx.
    Cot: SKU | Don vi tinh | SL | SL ban | SL KM
    Khong co cot COMBO -> Seller SKU chinh la SKU.
    Tra ve dict: {sku: {seller_sku, sku, unit, qty, qty_sold, promo_qty}}
    """
    from openpyxl import load_workbook

    wb = load_workbook(retail_path, data_only=True)
    ws = wb.active

    headers = {}
    for ci, cell in enumerate(ws[1], 1):
        val = str(cell.value).lower().strip() if cell.value else ''
        headers[ci] = val

    def find_col(keywords):
        for ci, h in headers.items():
            for kw in keywords:
                if kw in h:
                    return ci
        return 0

    col_sku = find_col(['sku', 'mã sản phẩm', 'ma san pham', 'mã sản phẩm']) or 1
    col_name = find_col(['tên sản phẩm', 'ten san pham', 'tên sp', 'product name']) or 0
    col_unit = find_col(['đơn vị tính', 'don vi tinh', 'đơn vị', 'don vi', 'dvt', 'unit']) or 0
    col_qty = find_col(['sl', 'qty', 'số lượng', 'so luong']) or 3
    col_sold = find_col(['sl bán', 'sl ban', 'qty sold', 'bán', 'ban']) or 5
    col_promo = find_col(['sl km', 'promo qty', 'khuyến mại', 'khuyen mai', 'km']) or 6

    retail = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = str(row[col_sku - 1]).strip() if row[col_sku - 1] is not None else ''
        if not sku:
            continue
        retail[sku] = {
            "seller_sku": sku,
            "sku": sku,
            "product_name": str(row[col_name - 1]).strip() if col_name and row[col_name - 1] is not None else '',
            "unit": str(row[col_unit - 1]).strip() if col_unit and row[col_unit - 1] is not None else '',
            "qty": int(row[col_qty - 1]) if row[col_qty - 1] is not None else 1,
            "qty_sold": int(row[col_sold - 1]) if row[col_sold - 1] is not None else 1,
            "promo_qty": int(row[col_promo - 1]) if row[col_promo - 1] is not None else 0,
        }
    wb.close()
    print(f"   Retail data: {len(retail)} SKUs")
    return retail


def build_sku_index(master_data: list[dict]) -> tuple[set[str], dict[str, str], dict[str, list[str]]]:
    """
    Từ master_data trả về:
      - master_skus: set tất cả Seller SKU
      - prefix_map: map từ SKU bị ngắt dòng → full Seller SKU (chỉ khi unique)
        VD: "CER01-" → "CER01-BOG17-2"
      - ambiguous_map: map từ prefix ambiguous → list các full SKU có thể
        VD: "ANTQ-3-" → ["ANTQ-3-6MET01", "ANTQ-3-TRG01"]
    """
    master_skus = set(r["seller_sku"] for r in master_data)

    prefix_map: dict[str, str] = {}
    ambiguous_map: dict[str, list[str]] = {}
    for sku in master_skus:
        if '-' in sku:
            parts = sku.split('-')
            for i in range(1, len(parts)):
                prefix = '-'.join(parts[:i]) + '-'
                if prefix not in prefix_map and prefix not in ambiguous_map:
                    prefix_map[prefix] = sku
                elif prefix in prefix_map:
                    # Trở thành ambiguous
                    ambiguous_map[prefix] = [prefix_map[prefix], sku]
                    del prefix_map[prefix]
                else:
                    ambiguous_map[prefix].append(sku)

    return master_skus, prefix_map, ambiguous_map


# ============================================================
# PDF EXTRACTION
# ============================================================

def extract_order_counts(
    pdf_path: str,
    master_skus: set[str],
    prefix_map: dict[str, str],
    ambiguous_map: dict[str, list[str]],
    retail_lookup: dict[str, dict] | None = None,
) -> dict[str, int]:
    """
    Trich xuat so don hang cho moi Seller SKU tu PDF.
    Dung regex tim pattern: SellerSKU + Qty + OrderID (15+ chu so).

    Thu tu uu tien:
      1. Exact match trong master_skus (combo)
      2. Prefix match unique
      3. Prefix ambiguous -> doan tu context
      4. Tim trong retail_lookup (san pham don le)
      5. Khong tim thay -> bo qua + canh bao

    Tra ve: {seller_sku: tong_so_don_hang}
    """
    with pdfplumber.open(pdf_path) as pdf:
        texts = []
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                texts.append(t)
    full_text = "\n".join(texts)

    # Gộp các dòng để xử lý SKU ngắt dòng (thay \n = space)
    flat_text = full_text.replace('\n', ' ')

    # Pattern: Mã SKU (chứa ít nhất 1 chữ cái, có thể bắt đầu bằng số, có thể kết thúc bằng - nếu bị ngắt)
    #           + Qty (số đơn hàng) + OrderID (15+ chữ số)
    pattern = r'\b((?=[A-Za-z0-9]*[A-Za-z])[A-Za-z0-9]+(?:-[A-Za-z0-9]+)*(?:-)?)\s+(\d+)\s+(\d{15,})'

    counts: dict[str, int] = defaultdict(int)

    # Từ noise thường bị regex bắt nhầm thành SKU
    NOISE_WORDS = {'combo', 'comboo', 'khong', 'kèm', 'quà', 'tặng', 'tặng)', 'cho', 'và',
                   'image', 'product', 'picking', 'list', 'order', 'item', 'user', 'print'}

    for match in re.finditer(pattern, flat_text):
        candidate = match.group(1)
        qty = int(match.group(2))

        if candidate in master_skus:
            # Exact match: Seller SKU xuat hien day du trong PDF
            counts[candidate] += qty
        elif candidate.endswith('-') and candidate in prefix_map:
            # Partial match unique: SKU bi ngat dong, tra prefix_map
            full_sku = prefix_map[candidate]
            counts[full_sku] += qty
        elif candidate.endswith('-') and candidate in ambiguous_map:
            # Prefix ambiguous: thu doan tu context sau Order ID
            options = ambiguous_map[candidate]
            after_text = flat_text[match.end():match.end()+120]
            best_match = None
            for opt in options:
                suffix = opt[len(candidate):]
                if suffix and suffix in after_text:
                    best_match = opt
                    break
            if best_match:
                counts[best_match] += qty
        elif candidate.endswith('-'):
            # Prefix khong co trong map nao -> bo qua
            pass
        elif retail_lookup and candidate in retail_lookup:
            # Tim thay trong san pham ban le
            counts[candidate] += qty
        else:
            # SKU la - khong co trong combo lan retail
            print(f"   ⚠ SKU la: {candidate} (x{qty}) - khong co trong ca 2 file")

    return dict(counts)


# ============================================================
# CALCULATION
# ============================================================

def calculate_results(
    master_data: list[dict],
    order_counts: dict[str, int],
    retail_lookup: dict[str, dict] | None = None,
) -> list[dict]:
    """
    Doi chieu master_data voi order_counts tu PDF.
    - Seller SKU co trong master_data -> combo (theo dinh luong)
    - Seller SKU co trong retail_lookup -> san pham ban le
    - Khong tim thay -> canh bao + bo qua
    - Seller SKU không có trong master_data → sản phẩm đơn lẻ (1 đơn = 1 sp)
    """
    results = []
    matched_skus = set()
    for row in master_data:
        seller_sku = row["seller_sku"]
        if seller_sku in order_counts:
            matched_skus.add(seller_sku)
            mult = order_counts[seller_sku]
            # Tra cứu tên sản phẩm từ file retail (nếu SKU có trong đó)
            product_name = ''
            if retail_lookup:
                ri = retail_lookup.get(row["sku"])
                if ri:
                    product_name = ri.get("product_name", '')
            results.append({
                "seller_sku": seller_sku,
                "sku": row["sku"],
                "qty": row["qty"] * mult,
                "qty_sold": row["qty_sold"] * mult,
                "promo_qty": row["promo_qty"] * mult,
                "unit": row["unit"],
                "product_name": product_name,
            })

    # Sản phẩm bán lẻ: có trong PDF nhưng không có trong master_data combo
    for seller_sku, count in order_counts.items():
        if seller_sku not in matched_skus and retail_lookup and seller_sku in retail_lookup:
            r = retail_lookup[seller_sku]
            results.append({
                "seller_sku": seller_sku,
                "sku": r["sku"],
                "qty": r["qty"] * count,
                "qty_sold": r["qty_sold"] * count,
                "promo_qty": r["promo_qty"] * count,
                "unit": r["unit"],
                "product_name": r.get("product_name", ''),
            })
            matched_skus.add(seller_sku)

    # Cảnh báo SKU không tìm thấy ở đâu
    for seller_sku in order_counts:
        if seller_sku not in matched_skus:
            print(f"   ⚠ SKU không xác định: {seller_sku} (x{order_counts[seller_sku]}) - bỏ qua")

    return results


# ============================================================
# EXCEL GENERATION
# ============================================================

def generate_excel(results: list[dict], output_path: str) -> str:
    """
    Tạo file Excel kết quả.
    Format: Seller SKU | SKU | Qty | Qty Sold | Promo Qty
    Dòng cuối: Tổng + SUM formula.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # ── Styles ──
    hdr_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    data_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", size=10, bold=True)

    headers = ["STT", "Seller SKU", "SKU", "Tên sản phẩm", "Qty", "Qty Sold", "Promo Qty"]
    col_widths = [5, 16, 12, 32, 7, 7, 7]  # tổng ~86 vừa A4 ngang
    col_aligns = ['C', 'C', 'C', 'L', 'R', 'R', 'R']

    # ── Header ──
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        c.border = thin_border

    # ── Data ──
    for ri, r in enumerate(results):
        rn = ri + 2
        vals = [ri + 1, r["seller_sku"], r["sku"], r.get("product_name", ""), r["qty"], r["qty_sold"], r["promo_qty"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=rn, column=ci, value=v)
            c.font = data_font
            c.border = thin_border
            ha = col_aligns[ci - 1] if ci <= len(col_aligns) else 'L'
            c.alignment = Alignment(horizontal={'C':'center','L':'left','R':'right'}.get(ha, 'left'),
                                    vertical='center',
                                    wrap_text=(ci == 4))  # wrap cột Tên sản phẩm

    # ── Total row ──
    tr = len(results) + 2
    tong_qty = sum(r["qty"] for r in results)
    tong_sold = sum(r["qty_sold"] for r in results)
    tong_promo = sum(r["promo_qty"] for r in results)

    ws.cell(row=tr, column=1, value="Tổng").font = total_font
    ws.cell(row=tr, column=1).border = thin_border
    ws.cell(row=tr, column=2).border = thin_border  # Seller SKU trống
    ws.cell(row=tr, column=3).border = thin_border  # SKU trống
    ws.cell(row=tr, column=4).border = thin_border  # Tên sản phẩm trống

    for ci, val in [(5, tong_qty), (6, tong_sold), (7, tong_promo)]:
        c = ws.cell(row=tr, column=ci, value=val)
        c.font = total_font
        c.border = thin_border

    # ── Column widths ──
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Print setup: vừa trang in, tránh mất cột ──
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = 9  # A4

    wb.save(output_path)
    print(f"   📊 Đã tạo Excel: {os.path.basename(output_path)}")
    return output_path


def generate_grouped_excel(results: list[dict], output_path: str, carrier: str = '', source_label: str = '') -> str:
    """
    Tạo file Excel gộp theo SKU (không hiện Seller SKU).
    Format: SKU | Đơn vị tính | Qty | Qty Sold | Promo Qty
    Có dòng tiêu đề in đậm ở đầu để nhận diện khi in giấy.
    """
    # Gộp theo SKU (giống logic generate_grouped_pdf)
    grouped = {}
    for r in results:
        sku = r["sku"]
        if sku not in grouped:
            grouped[sku] = {"qty": 0, "qty_sold": 0, "promo_qty": 0, "unit": r.get("unit", ""), "product_name": r.get("product_name", "")}
        grouped[sku]["qty"] += r["qty"]
        grouped[sku]["qty_sold"] += r["qty_sold"]
        grouped[sku]["promo_qty"] += r["promo_qty"]
        # Giữ product_name đầu tiên khác rỗng
        if not grouped[sku]["product_name"] and r.get("product_name", ""):
            grouped[sku]["product_name"] = r["product_name"]

    grouped_list = [{"sku": k, **v} for k, v in grouped.items()]
    grouped_list.sort(key=lambda x: x["sku"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # ── Styles ──
    title_font = Font(name="Arial", size=14, bold=True, color="1F4E79")
    title_align = Alignment(horizontal="center", vertical="center")
    hdr_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    data_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", size=10, bold=True)

    # ── Title row (dòng nhận diện khi in giấy) ──
    ncols = 7  # STT, SKU, Tên SP, ĐVT, SL, SL bán, SL KM
    title_text = f'{carrier} — {source_label}' if carrier and source_label else (carrier or source_label or 'Báo cáo gộp SKU')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title_text)
    c.font = title_font
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28  # đủ cao để hiển thị rõ

    # ── Column headers (row 2) ──
    headers = ["STT", "SKU", "Tên sản phẩm", "Đơn vị tính", "SL", "SL bán", "SL KM"]
    # Độ rộng cột tối ưu: SKU=12, Tên SP=36 wrap, cột số=7 → tổng ~86 vừa A4 ngang
    col_widths = [5, 12, 36, 12, 7, 7, 7]
    col_aligns = ['C', 'C', 'L', 'C', 'R', 'R', 'R']  # center / left / right

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border

    # ── Data (bắt đầu từ row 3) ──
    for ri, r in enumerate(grouped_list):
        rn = ri + 3
        vals = [ri + 1, r["sku"], r.get("product_name", ""), r.get("unit", ""), r["qty"], r["qty_sold"], r["promo_qty"]]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=rn, column=ci, value=v)
            c.font = data_font
            c.border = thin_border
            # Căn lề theo cột + wrap text cho cột Tên sản phẩm (cột 3)
            ha = col_aligns[ci - 1] if ci <= len(col_aligns) else 'L'
            c.alignment = Alignment(horizontal={'C':'center','L':'left','R':'right'}.get(ha, 'left'),
                                    vertical='center',
                                    wrap_text=(ci == 3))

    # ── Total row ──
    tr = len(grouped_list) + 3
    tong_qty = sum(r["qty"] for r in grouped_list)
    tong_sold = sum(r["qty_sold"] for r in grouped_list)
    tong_promo = sum(r["promo_qty"] for r in grouped_list)

    ws.cell(row=tr, column=1, value="Tổng").font = total_font
    ws.cell(row=tr, column=1).border = thin_border
    ws.cell(row=tr, column=2).border = thin_border  # SKU trống
    ws.cell(row=tr, column=3).border = thin_border  # Tên sản phẩm trống
    ws.cell(row=tr, column=4).border = thin_border  # Đơn vị tính trống

    for ci, val in [(5, tong_qty), (6, tong_sold), (7, tong_promo)]:
        c = ws.cell(row=tr, column=ci, value=val)
        c.font = total_font
        c.border = thin_border

    # ── Column widths ──
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Print setup: vừa trang in, tránh mất cột ──
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # tự động số dòng
    ws.page_setup.paperSize = 9  # A4

    wb.save(output_path)
    print(f"   📊 Đã tạo Excel gộp: {os.path.basename(output_path)}")
    return output_path


# ============================================================
# FONT HELPER
# ============================================================

def _find_font(bold: bool = False) -> tuple[str, str]:
    """Tìm font Arial trên Windows, fallback về Helvetica (FPDF built-in)."""
    if bold:
        paths = [r"C:\Windows\Fonts\Arialbd.ttf", r"C:\Windows\Fonts\Arial Bold.ttf"]
    else:
        paths = [r"C:\Windows\Fonts\Arial.ttf", r"C:\Windows\Fonts\Arial Regular.ttf"]
    for p in paths:
        if os.path.exists(p):
            return "Arial", p
    return "Helvetica", None


# ============================================================
# PDF GENERATION
# ============================================================

def generate_pdf(results: list[dict], output_path: str) -> str:
    """Tạo file PDF báo cáo đã tách SKU."""
    print(f"   📝 Tạo PDF: {os.path.basename(output_path)}")

    pdf = FPDF(orientation='P', unit='pt', format='Letter')
    pdf.set_auto_page_break(auto=True, margin=40)

    # Dùng font Arial TTF với unicode (hỗ trợ tiếng Việt)
    font_path = r"C:\Windows\Fonts\Arial.ttf"
    font_bold_path = r"C:\Windows\Fonts\Arialbd.ttf"
    font_name = 'Arial'
    font_bold = 'ArialBold'

    if os.path.exists(font_path):
        pdf.add_font(font_name, '', font_path, uni=True)
    else:
        font_name = 'Helvetica'

    if os.path.exists(font_bold_path):
        pdf.add_font(font_bold, 'B', font_bold_path, uni=True)
    else:
        font_bold = 'Helvetica'

    pdf.add_page()

    GRAY, BLACK = (100, 100, 100), (0, 0, 0)

    # Header
    pdf.set_font(font_name, "", 8)
    pdf.set_text_color(*GRAY)
    pdf.set_xy(0, 16)
    pdf.cell(612, 10, "                                     Traphaco ~ Danh sách đơn hàng ~ Haravan", align="C")

    y = 37
    pdf.set_text_color(*BLACK)
    pdf.set_font(font_bold, "B", 13.5)
    pdf.set_xy(28.5, y)
    pdf.cell(0, 18, "Traphaco")
    y += 18

    pdf.set_font(font_name, "", 9)
    for line in [
        f"Ngày in phiếu: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
        "Người in: Trường Lê",
        "Địa chỉ: Ngõ 15 Ngọc Hồi, Phường Hoàng Liệt, Quận Hoàng Mai, Hà Nội, Vietnam",
        "Điện thoại: 0989821222",
    ]:
        pdf.set_xy(28.5, y)
        pdf.cell(0, 12, line)
        y += 12
    y += 21

    pdf.set_font(font_bold, "B", 21)
    pdf.set_xy(0, y)
    pdf.cell(612, 28, "Danh sách sản phẩm", align="C")
    y += 60

    # ── Table ──
    ROW_H = 22
    RIGHT = 583.5
    # (label, x, width, align) — align áp dụng cho cả header & data
    col_defs = [
        ("STT",        28.5,  30, "R"),
        ("Seller SKU", 66.5, 135, "L"),
        ("SKU",       209.5, 100, "L"),
        ("Qty",       317.5,  80, "R"),
        ("Qty Sold",  405.5,  85, "R"),
        ("Promo Qty", 498.5,  85, "R"),
    ]

    def draw_header(yy):
        pdf.set_draw_color(*BLACK)
        pdf.set_line_width(0.5)
        pdf.line(28.5, yy, RIGHT, yy)
        pdf.set_font(font_bold, "B", 8)
        for label, x, w, align in col_defs:
            pdf.set_xy(x, yy + 2)
            pdf.cell(w, ROW_H, label, align=align)
        return yy + ROW_H

    y = draw_header(y)
    pdf.line(28.5, y, RIGHT, y)

    pdf.set_font(font_name, "", 8)
    tong_qty = tong_sold = tong_promo = 0

    for i, r in enumerate(results, 1):
        if y > 720:
            pdf.line(28.5, y, RIGHT, y)
            pdf.add_page()
            y = 40
            y = draw_header(y)
            pdf.line(28.5, y, RIGHT, y)
            pdf.set_font(font_name, "", 8)

        row_vals = [str(i), r["seller_sku"], r["sku"],
                    str(r["qty"]), str(r["qty_sold"]), str(r["promo_qty"])]
        for (_, x, w, align), v in zip(col_defs, row_vals):
            pdf.set_xy(x, y + 2)
            pdf.cell(w, ROW_H, v, align=align)

        tong_qty += r["qty"]
        tong_sold += r["qty_sold"]
        tong_promo += r["promo_qty"]
        y += ROW_H

    pdf.line(28.5, y, RIGHT, y)
    y += 6
    pdf.set_font(font_bold, "B", 8)

    total_cells = [
        ("Tổng", col_defs[2][1], col_defs[2][2], "L"),
        (str(tong_qty), col_defs[3][1], col_defs[3][2], "R"),
        (str(tong_sold), col_defs[4][1], col_defs[4][2], "R"),
        (str(tong_promo), col_defs[5][1], col_defs[5][2], "R"),
    ]
    for txt, x, w, align in total_cells:
        pdf.set_xy(x, y)
        pdf.cell(w, ROW_H, txt, align=align)

    # Footer
    pdf.set_font(font_name, "", 8)
    pdf.set_text_color(*GRAY)
    pdf.set_xy(0, y + 30)
    pdf.cell(612, 10, "https://traphaco.myharavan.com/admin/orders                                                    1/1", align="C")

    pdf.output(output_path)
    return output_path


# ============================================================
# TXT GENERATION
# ============================================================

def generate_txt(results: list[dict], output_path: str) -> str:
    """Tạo file TXT báo cáo."""
    print(f"   📄 Tạo TXT: {os.path.basename(output_path)}")

    tong_qty = sum(r["qty"] for r in results)
    tong_sold = sum(r["qty_sold"] for r in results)
    tong_promo = sum(r["promo_qty"] for r in results)

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("Danh sách sản phẩm đã đối chiếu với Master Data\n")
        f.write(f"Thời gian: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n")
        f.write(f"Số dòng: {len(results)}\n")
        f.write(f"Tổng Qty: {tong_qty}  |  Qty Sold: {tong_sold}  |  Promo Qty: {tong_promo}\n")
        f.write(f"{'=' * 75}\n")
        f.write(f"{'STT':<6} {'Seller SKU':<18} {'SKU':<14} {'Qty':>6} {'Sold':>6} {'Promo':>6}\n")
        f.write(f"{'-' * 75}\n")
        for i, r in enumerate(results, 1):
            f.write(f"{i:<6} {r['seller_sku']:<18} {r['sku']:<14} "
                    f"{r['qty']:>6} {r['qty_sold']:>6} {r['promo_qty']:>6}\n")
    return output_path


# ============================================================
# GROUPED PDF (gộp theo SKU, bỏ cột Seller SKU)
# ============================================================

def generate_grouped_pdf(results: list[dict], output_path: str) -> str:
    """Tạo PDF báo cáo đã gộp theo SKU (không hiện Seller SKU)."""
    print(f"   📝 Tạo PDF gộp: {os.path.basename(output_path)}")

    # Gộp theo SKU
    grouped = {}
    for r in results:
        sku = r["sku"]
        if sku not in grouped:
            grouped[sku] = {"qty": 0, "qty_sold": 0, "promo_qty": 0}
        grouped[sku]["qty"] += r["qty"]
        grouped[sku]["qty_sold"] += r["qty_sold"]
        grouped[sku]["promo_qty"] += r["promo_qty"]

    grouped_list = [{"sku": k, **v} for k, v in grouped.items()]
    # Sắp xếp theo SKU
    grouped_list.sort(key=lambda x: x["sku"])

    pdf = FPDF(orientation='P', unit='pt', format='Letter')
    pdf.set_auto_page_break(auto=True, margin=40)

    font_path = r"C:\Windows\Fonts\Arial.ttf"
    font_bold_path = r"C:\Windows\Fonts\Arialbd.ttf"
    font_name = 'Arial'
    font_bold = 'ArialBold'

    if os.path.exists(font_path):
        pdf.add_font(font_name, '', font_path, uni=True)
    else:
        font_name = 'Helvetica'

    if os.path.exists(font_bold_path):
        pdf.add_font(font_bold, 'B', font_bold_path, uni=True)
    else:
        font_bold = 'Helvetica'

    pdf.add_page()

    GRAY, BLACK = (100, 100, 100), (0, 0, 0)

    # Header
    pdf.set_font(font_name, "", 8)
    pdf.set_text_color(*GRAY)
    pdf.set_xy(0, 16)
    pdf.cell(612, 10, "                                     Traphaco ~ Danh sách đơn hàng ~ Haravan", align="C")

    y = 37
    pdf.set_text_color(*BLACK)
    pdf.set_font(font_bold, "B", 13.5)
    pdf.set_xy(28.5, y); pdf.cell(0, 18, "Traphaco")
    y += 18

    pdf.set_font(font_name, "", 9)
    for line in [
        f"Ngày in phiếu: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
        "Người in: Trường Lê",
        "Địa chỉ: Ngõ 15 Ngọc Hồi, P.Hoàng Liệt, Q.Hoàng Mai, Hà Nội",
        "Điện thoại: 0989821222",
    ]:
        pdf.set_xy(28.5, y); pdf.cell(0, 12, line)
        y += 12
    y += 21

    pdf.set_font(font_bold, "B", 16)
    pdf.set_xy(0, y); pdf.cell(612, 28, "Danh sách sản phẩm (gộp theo SKU)", align="C")
    y += 60

    # ── Table ──
    ROW_H = 24
    RIGHT = 583.5
    col_defs = [
        ("STT",    28.5,  40, "R"),
        ("SKU",    76.5, 160, "L"),
        ("Qty",   244.5, 100, "R"),
        ("Qty Sold", 352.5, 110, "R"),
        ("Promo Qty", 470.5, 110, "R"),
    ]

    def draw_header(yy):
        pdf.set_draw_color(*BLACK); pdf.set_line_width(0.5)
        pdf.line(28.5, yy, RIGHT, yy)
        pdf.set_font(font_bold, "B", 9)
        for label, x, w, align in col_defs:
            pdf.set_xy(x, yy + 2); pdf.cell(w, ROW_H, label, align=align)
        return yy + ROW_H

    y = draw_header(y)
    pdf.line(28.5, y, RIGHT, y)
    pdf.set_font(font_name, "", 9)
    tong_qty = tong_sold = tong_promo = 0

    for i, r in enumerate(grouped_list, 1):
        if y > 720:
            pdf.line(28.5, y, RIGHT, y); pdf.add_page(); y = 40
            y = draw_header(y)
            pdf.line(28.5, y, RIGHT, y); pdf.set_font(font_name, "", 9)

        row_vals = [str(i), r["sku"], str(r["qty"]), str(r["qty_sold"]), str(r["promo_qty"])]
        for (_, x, w, align), v in zip(col_defs, row_vals):
            pdf.set_xy(x, y + 2); pdf.cell(w, ROW_H, v, align=align)

        tong_qty += r["qty"]; tong_sold += r["qty_sold"]; tong_promo += r["promo_qty"]
        y += ROW_H

    pdf.line(28.5, y, RIGHT, y); y += 6
    pdf.set_font(font_bold, "B", 9)

    total_cells = [
        ("Tổng", col_defs[1][1], col_defs[1][2], "L"),
        (str(tong_qty), col_defs[2][1], col_defs[2][2], "R"),
        (str(tong_sold), col_defs[3][1], col_defs[3][2], "R"),
        (str(tong_promo), col_defs[4][1], col_defs[4][2], "R"),
    ]
    for txt, x, w, align in total_cells:
        pdf.set_xy(x, y); pdf.cell(w, ROW_H, txt, align=align)

    pdf.set_font(font_name, "", 8)
    pdf.set_text_color(*GRAY)
    pdf.set_xy(0, y + 30)
    pdf.cell(612, 10, "https://traphaco.myharavan.com/admin/orders                                                    1/1", align="C")

    pdf.output(output_path)
    return output_path


# ============================================================
# SINGLE PDF PIPELINE
# ============================================================

def process_single_pdf(
    pdf_path: str,
    output_dir: str,
    master_data: list[dict],
    master_skus: set[str],
    prefix_map: dict[str, str],
    ambiguous_map: dict[str, list[str]],
    retail_lookup: dict[str, dict] | None = None,
    carrier: str = '',
) -> dict | None:
    """
    Xu ly 1 file PDF: trich xuat -> doi chieu -> xuat file.
    Tra ve dict thong tin ket qua, hoac None neu khong co du lieu.
    """
    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
    print(f"\nXu ly: {os.path.basename(pdf_path)}")

    # B1: Trich xuat
    order_counts = extract_order_counts(pdf_path, master_skus, prefix_map, ambiguous_map, retail_lookup)
    if not order_counts:
        print(f"   ⚠ Khong tim thay Seller SKU nao trong PDF!")
        return None

    print(f"   Tim thay {len(order_counts)} Seller SKU, "
          f"tong {sum(order_counts.values())} don hang")

    # B2: Tinh toan
    results = calculate_results(master_data, order_counts, retail_lookup)
    print(f"   📊 Kết quả: {len(results)} dòng")

    # B3: Xuất file
    # B3: Xuất file — tên rõ ràng, phân biệt với PDF gốc
    # Tên file có gắn tên carrier để phân biệt khi có nhiều đơn vị vận chuyển
    carrier_safe = carrier.replace(' ', '_').replace('&', 'n') if carrier else ''

    # Bỏ carrier prefix khỏi base_name nếu đã có (tránh lặp)
    clean_name = base_name
    if carrier_safe and base_name.startswith(carrier_safe + '_'):
        clean_name = base_name[len(carrier_safe) + 1:]

    # Trích timestamp: hỗ trợ cả định dạng TikTok (MM-DD_HH-MM-SS) và định dạng cũ (YYYYMMDD_HHMMSS)
    ts_match = re.search(r'(\d{2}-\d{2}_\d{2}-\d{2}-\d{2}|\d{8}_\d{6})', clean_name)
    ts = ts_match.group(1) if ts_match else clean_name

    prefix = f"Bao_cao_gop_SKU_{carrier_safe}_" if carrier_safe else "Bao_cao_gop_SKU_"
    excel_grouped_path = os.path.join(output_dir, f"{prefix}{ts}.xlsx")

    # Tạo label nhận diện cho dòng tiêu đề trong Excel (hiển thị khi in giấy)
    # clean_name VD: "06-30_14-11-10_Picking list_1"
    # Chuyển thành: "Picking list 1 — 30/06 14:11"
    label_match = re.search(r'(\d{2})-(\d{2})_(\d{2})-(\d{2})-(\d{2})_(.+)', clean_name)
    if label_match:
        mm, dd, hh, mi, ss = label_match.group(1), label_match.group(2), label_match.group(3), label_match.group(4), label_match.group(5)
        remainder = label_match.group(6).replace('_', ' ')
        source_label = f'{remainder} — {dd}/{mm} {hh}:{mi}'
    else:
        source_label = clean_name.replace('_', ' ')

    # Chỉ xuất file Excel gộp theo SKU (không cột Seller SKU)
    # generate_pdf(results, pdf_out_path)
    # generate_grouped_pdf(results, pdf_grouped_path)
    generate_grouped_excel(results, excel_grouped_path, carrier=carrier, source_label=source_label)

    return {
        "base_name": base_name,
        "rows": len(results),
        "tong_qty": sum(r["qty"] for r in results),
        "tong_sold": sum(r["qty_sold"] for r in results),
        "tong_promo": sum(r["promo_qty"] for r in results),
        "files": {
            "excel": excel_grouped_path,
        },
    }


# ============================================================
# MAIN ENTRY POINT
# ============================================================

def process_all(
    pdf_files: list[str],
    output_dir: str,
    master_path: str | None = None,
    retail_path: str | None = None,
    carrier: str = '',
) -> list[dict]:
    """
    Xu ly toan bo pipeline cho nhieu file PDF.
    Moi PDF duoc doi chieu doc lap voi master_data va sinh file ket qua rieng.

    Args:
        pdf_files: danh sach duong dan file PDF
        output_dir: thu muc xuat ket qua
        master_path: duong dan ma combo.xlsx (combo)
        retail_path: duong dan sp ban le.xlsx (don le)
    """
    # Xac dinh master_path
    if master_path is None:
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        master_path = os.path.join(base, "master_data.xlsx")

    if not os.path.exists(master_path):
        raise ValueError(f"Khong tim thay file master_data tai: {master_path}")

    # Load master data (combo)
    print(f"Load master data: {master_path}")
    master_data = load_master_data(master_path)
    master_skus, prefix_map, ambiguous_map = build_sku_index(master_data)
    print(f"   {len(master_data)} dong, {len(master_skus)} Seller SKU"
          + (f", {len(prefix_map)} prefix map" if prefix_map else ""))

    # Load retail data (san pham ban le)
    retail_lookup = None
    if retail_path and os.path.exists(retail_path):
        retail_lookup = load_retail_data(retail_path)
    elif retail_path:
        print(f"   ⚠ Khong tim thay file retail: {retail_path}")

    # Xu ly tung PDF
    all_results = []
    for pdf_path in pdf_files:
        result = process_single_pdf(pdf_path, output_dir,
                                    master_data, master_skus, prefix_map, ambiguous_map, retail_lookup, carrier)
        if result:
            all_results.append(result)

    if not all_results:
        raise ValueError("Không trích xuất được dữ liệu từ bất kỳ PDF nào. "
                         "Kiểm tra file đầu vào và master_data.")

    return all_results


# ============================================================
# CLI — Test nhanh
# ============================================================
if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python calculator.py <pdf_file> [pdf_file2...]")
        sys.exit(1)

    pdfs = sys.argv[1:]
    out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "outputs")
    os.makedirs(out_dir, exist_ok=True)

    master = os.path.join(os.path.dirname(os.path.abspath(__file__)), "master_data.xlsx")

    try:
        results = process_all(pdfs, out_dir, master)
        print("\n" + "=" * 60)
        print("✅ HOÀN THÀNH!")
        for r in results:
            print(f"\n📦 {r['base_name']}:")
            print(f"   Dòng: {r['rows']} | Qty: {r['tong_qty']} | "
                  f"Sold: {r['tong_sold']} | Promo: {r['tong_promo']}")
            for k, v in r["files"].items():
                print(f"   📎 {k}: {os.path.basename(v)}")
    except ValueError as e:
        print(f"\n❌ Lỗi: {e}")
        sys.exit(1)
